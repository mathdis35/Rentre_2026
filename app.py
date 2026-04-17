import os, re, shutil, datetime, uuid, json, copy, calendar
from pathlib import Path
from collections import defaultdict
from flask import Flask, request, jsonify, send_file, after_this_request
from flask_cors import CORS

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    import xlrd
except ImportError:
    pass

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response
UPLOAD_FOLDER = '/tmp/plannipro'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MOIS_ABBR = {
    'août':8,'aout':8,'sept':9,'septembre':9,'oct':10,'octobre':10,
    'nov':11,'novembre':11,'déc':12,'dec':12,'décembre':12,
    'janv':1,'jan':1,'janvier':1,'févr':2,'fev':2,'février':2,
    'mars':3,'avr':4,'avril':4,'mai':5,'juin':6,'juil':7,'juillet':7
}
MOIS_FR = ['','Janvier','Février','Mars','Avril','Mai','Juin',
           'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
MOIS_FR_UPPER = {
    1:'JANVIER',2:'FÉVRIER',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
    7:'JUILLET',8:'AOÛT',9:'SEPTEMBRE',10:'OCTOBRE',11:'NOVEMBRE',12:'DÉCEMBRE'
}

FERIES = {
    datetime.date(2026,11,1), datetime.date(2026,11,11), datetime.date(2026,12,25),
    datetime.date(2027,1,1),  datetime.date(2027,4,5),   datetime.date(2027,5,1),
    datetime.date(2027,5,8),  datetime.date(2027,5,13),  datetime.date(2027,5,24),
    datetime.date(2027,7,14), datetime.date(2027,8,15),
}

DEFAULT_COLORS = ['FFEE7E32','FFFFCC99','FFFFD243','FFDDFFDD',
                  'FFB8D4F0','FFFFC0CB','FFD4B8F0','FFB8F0D4']


# ─── Paires de colonnes fusionnées dans le template ──────────────────────────
# (découvertes par analyse du template vierge — ne pas modifier)
ALL_MERGE_PAIRS = [
    (4,5),(7,8),(10,11),(13,14),(16,17),(19,20),
    (28,29),(31,32),(34,35),(37,38),(40,41),(43,44),(46,47),(49,50),
    (56,57),(59,60),(62,63),(65,66),(68,69),(71,72),
    (78,79),(81,82),(84,85),(87,88),(90,91)
]
JOURS_COLS   = [2, 26, 54, 76]
JOURS_FR_LIST = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi']

# ─── Helpers ──────────────────────────────────────────────────────────────────
def colors_match(rgb):
    if not rgb: return False
    return abs(rgb[0]-166)<20 and abs(rgb[1]-202)<20 and abs(rgb[2]-240)<20

def cell_to_str(v):
    if v is None: return ''
    return str(v).strip()

def is_available(v):
    if v is None: return False
    return str(v).strip().upper() in ['X','4','✓','OUI','O']

def find_month_num(text):
    t = text.strip().lower()
    for m, mn in MOIS_ABBR.items():
        if m in t:
            yr = re.search(r'(20\d\d)', t)
            return mn, int(yr.group(1)) if yr else None
    return None, None

# ─── Parser planning classe XLS ───────────────────────────────────────────────
def parse_planning_xls(filepath):
    try:
        wb = xlrd.open_workbook(filepath, formatting_info=True)
    except Exception as e:
        return {'nom': Path(filepath).stem, 'jours': []}
    ws = wb.sheet_by_index(0)
    nom = Path(filepath).stem
    for c in range(ws.ncols):
        v = ws.cell_value(0, c)
        if isinstance(v, str) and len(v.strip()) > 3:
            nom = v.strip(); break

    jours = []
    blocks = []
    for c in range(ws.ncols):
        v = ws.cell_value(4, c)
        if isinstance(v, str):
            mn, yr = find_month_num(v)
            if mn and yr:
                blocks.append({'cj': c+1, 'cd': c+2, 'y': yr, 'm': mn})

    for b in blocks:
        for r in range(5, ws.nrows):
            if b['cj'] >= ws.ncols or b['cd'] >= ws.ncols: continue
            try:
                xf = wb.xf_list[ws.cell_xf_index(r, b['cj'])]
                rgb = wb.colour_map.get(xf.background.pattern_colour_index)
                if not colors_match(rgb): continue
            except: continue
            dv = ws.cell_value(r, b['cd'])
            dn = int(dv) if isinstance(dv, float) and dv > 0 else None
            if isinstance(dv, str) and dv.strip().isdigit(): dn = int(dv.strip())
            if not dn or not (1 <= dn <= 31): continue
            try:
                d = datetime.date(b['y'], b['m'], dn)
                if d.weekday() < 5: jours.append(d.strftime('%Y-%m-%d'))
            except: pass
    return {'nom': nom, 'jours': sorted(set(jours))}

# ─── Parser planning classe XLSX ──────────────────────────────────────────────
def parse_planning_xlsx(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
    except: return {'nom': Path(filepath).stem, 'jours': []}
    ws = wb.active
    nom = Path(filepath).stem
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        if v and isinstance(v, str) and len(v.strip()) > 3:
            nom = v.strip(); break

    jours = []
    blocks = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=5, column=c).value
        if isinstance(v, str):
            mn, yr = find_month_num(v)
            if mn and yr:
                blocks.append({'cj': c+1, 'cd': c+2, 'y': yr, 'm': mn})

    for b in blocks:
        for r in range(6, ws.max_row+1):
            cell = ws.cell(row=r, column=b['cj'])
            bg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb' else None
            if bg not in ('FFA6CAF0', 'A6CAF0'): continue
            dv = ws.cell(row=r, column=b['cd']).value
            dn = None
            if isinstance(dv, datetime.datetime): dn = dv.day
            elif isinstance(dv, (int, float)): dn = int(dv)
            elif isinstance(dv, str) and dv.strip().isdigit(): dn = int(dv.strip())
            if not dn or not (1 <= dn <= 31): continue
            try:
                d = datetime.date(b['y'], b['m'], dn)
                if d.weekday() < 5: jours.append(d.strftime('%Y-%m-%d'))
            except: pass
    return {'nom': nom, 'jours': sorted(set(jours))}

def parse_planning_classe(fp):
    return parse_planning_xls(fp) if Path(fp).suffix.lower() == '.xls' else parse_planning_xlsx(fp)

# ─── Parser disponibilités ────────────────────────────────────────────────────
def parse_disponibilite(filepath):
    nom = Path(filepath).stem
    dispo = {}
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except:
        return {'nom': nom, 'dispo': {}}

    if len(rows) < 2: return {'nom': nom, 'dispo': {}}

    header = rows[0]
    month_cols = []
    for ci, val in enumerate(header):
        if isinstance(val, str):
            mn, yr = find_month_num(val)
            if mn and yr:
                month_cols.append({'ci': ci, 'm': mn, 'y': yr})

    ign = {'sam', 'dim', 'férié', 'ferie', 'férie', 'fériés', 'nan', 'none', ''}

    for mc in month_cols:
        ci, mn, yr = mc['ci'], mc['m'], mc['y']
        col_matin = ci
        col_pm    = ci + 1
        for ri in range(2, len(rows)):
            row = rows[ri]
            day_abbr = cell_to_str(row[0]).lower() if len(row) > 0 else ''
            if day_abbr in ign: continue
            day_num_raw = row[1] if len(row) > 1 else None
            dn = None
            if isinstance(day_num_raw, (int, float)) and day_num_raw == day_num_raw:
                dn = int(day_num_raw)
            elif isinstance(day_num_raw, str) and day_num_raw.strip().isdigit():
                dn = int(day_num_raw.strip())
            if not dn or not (1 <= dn <= 31): continue
            try:
                d = datetime.date(yr, mn, dn)
                if d.weekday() >= 5: continue
                ds = d.strftime('%Y-%m-%d')
                mv = row[col_matin] if col_matin < len(row) else None
                pv = row[col_pm]    if col_pm    < len(row) else None
                if cell_to_str(mv).lower() in ign: continue
                dispo[ds] = {'matin': is_available(mv), 'pm': is_available(pv)}
            except: pass

    return {'nom': nom, 'dispo': dispo}

# ─── Parser tableau formateurs ────────────────────────────────────────────────
def parse_tableau_formateurs(filepath):
    assignments = defaultdict(list)

    def process(rows):
        cur = None
        for row in rows:
            v0 = cell_to_str(row[0]) if row and row[0] is not None else ''
            if re.match(r'(BTS|BAC|CGC|NDRC|GPME|Master|RDC|RH|EC)\s', v0, re.I):
                cur = v0
            elif cur and v0 and len(row) > 1 and row[1]:
                mat = cell_to_str(row[1])
                heures_raw = row[-1] if row else 0
                try: h = float(str(heures_raw).split('+')[0].strip())
                except: h = 0
                if h > 0 and mat and mat.lower() not in ['nan', 'total', '']:
                    assignments[cur].append({'formateur': v0, 'matiere': mat, 'heures': h, 'heures_faites': 0})

    ext = Path(filepath).suffix.lower()
    if ext == '.xls':
        try:
            wb = xlrd.open_workbook(filepath)
            si = next((i for i, n in enumerate(wb.sheet_names()) if 'mois' in n.lower()), len(wb.sheet_names())-1)
            ws = wb.sheet_by_index(si)
            process([[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)])
        except: pass
    else:
        try:
            wb = load_workbook(filepath, data_only=True)
            sn = next((s for s in wb.sheetnames if 'mois' in s.lower()), wb.sheetnames[-1])
            process(list(wb[sn].iter_rows(values_only=True)))
        except: pass
    return dict(assignments)

# ─── Parseur tableau formateurs v2 (format tabulaire) ────────────────────────

def parse_tableau_formateurs_v2(filepath):
    """
    Parse le nouveau format d'affectations : 1 ligne = 1 affectation.

    Feuille cible : 'AFFECTATIONS' (ou première feuille si absente).
    En-tête auto-détecté dans les 10 premières lignes.

    Colonnes OBLIGATOIRES : CLASSE, FORMATEUR, MATIERE, HEURES_ANNEE
    Colonnes OPTIONNELLES : PRIORITE (défaut 1), NOTES (ignoré), ACTIF (défaut OUI)

    Retourne un dict compatible avec assigner() :
      { "BTS MCO 73": [
            {"formateur": "J.Christophe", "matiere": "Gestion",
             "heures": 100.0, "heures_faites": 0, "priorite": 1},
            ...
        ], ...
      }
    """
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        raise ValueError(f"Impossible d'ouvrir {Path(filepath).name} : {e}")

    ws = wb['AFFECTATIONS'] if 'AFFECTATIONS' in wb.sheetnames else wb.active

    # ── Détecter la ligne d'en-tête ─────────────────────────────────────────
    header_row = None
    col_map    = {}
    for r in range(1, min(ws.max_row + 1, 10)):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        upper    = [str(v).strip().upper() if v is not None else '' for v in row_vals]
        if 'CLASSE' in upper and 'FORMATEUR' in upper and 'MATIERE' in upper:
            header_row = r
            for i, label in enumerate(upper):
                if label:
                    col_map[label] = i + 1   # 1-indexé
            break

    if header_row is None:
        raise ValueError(
            "En-tête introuvable dans la feuille AFFECTATIONS. "
            "Colonnes attendues : CLASSE, FORMATEUR, MATIERE, HEURES_ANNEE"
        )

    # ── Valider les colonnes obligatoires ────────────────────────────────────
    manquantes = [col for col in ('CLASSE', 'FORMATEUR', 'MATIERE', 'HEURES_ANNEE')
                  if col not in col_map]
    if manquantes:
        raise ValueError(f"Colonnes obligatoires manquantes : {', '.join(manquantes)}")

    c_classe = col_map['CLASSE']
    c_form   = col_map['FORMATEUR']
    c_mat    = col_map['MATIERE']
    c_h      = col_map['HEURES_ANNEE']
    c_prio   = col_map.get('PRIORITE')
    c_actif  = col_map.get('ACTIF')

    # ── Lire les affectations ────────────────────────────────────────────────
    assignments = defaultdict(list)
    lignes_ignorees = 0

    for r in range(header_row + 1, ws.max_row + 1):
        classe    = ws.cell(r, c_classe).value
        formateur = ws.cell(r, c_form).value
        matiere   = ws.cell(r, c_mat).value
        heures_v  = ws.cell(r, c_h).value
        actif_v   = ws.cell(r, c_actif).value if c_actif else 'OUI'
        prio_v    = ws.cell(r, c_prio).value  if c_prio  else 1

        # Ignorer lignes vides
        if not (classe and formateur and matiere):
            continue

        # Ignorer lignes ACTIF = NON
        if str(actif_v or 'OUI').strip().upper() == 'NON':
            lignes_ignorees += 1
            continue

        # Valider et normaliser les heures (obligatoires, > 0)
        try:
            heures = float(str(heures_v).replace(',', '.').strip())
        except (TypeError, ValueError):
            lignes_ignorees += 1
            continue
        if heures <= 0:
            lignes_ignorees += 1
            continue

        # Normaliser la priorité
        try:
            priorite = int(prio_v) if prio_v else 1
        except (TypeError, ValueError):
            priorite = 1

        assignments[str(classe).strip()].append({
            'formateur':     str(formateur).strip(),
            'matiere':       str(matiere).strip(),
            'heures':        heures,
            'heures_faites': 0,        # requis par assigner()
            'priorite':      priorite, # pour tri — assigner() l'ignore
        })

    # Trier chaque classe par priorité (1 = principal avant 2 = remplaçant)
    for classe in assignments:
        assignments[classe].sort(key=lambda x: x['priorite'])

    # ── Log debug ────────────────────────────────────────────────────────────
    nb_classes     = len(assignments)
    nb_affectations = sum(len(v) for v in assignments.values())
    nb_formateurs  = len({a['formateur'] for v in assignments.values() for a in v})
    import logging
    logging.info(
        f"[parse_v2] {nb_classes} classes | {nb_formateurs} formateurs | "
        f"{nb_affectations} affectations | {lignes_ignorees} lignes ignorées"
    )
    # Exposé aussi en attribut pour que la route puisse le retourner si besoin
    result = dict(assignments)
    result['_debug'] = {
        'classes':      nb_classes,
        'formateurs':   nb_formateurs,
        'affectations': nb_affectations,
        'ignores':      lignes_ignorees,
    }
    return result


def _auto_parse_formateurs(filepath):
    """
    Sélecteur automatique : détecte le format du fichier et appelle
    le bon parseur.

    - Feuille AFFECTATIONS présente → parse_tableau_formateurs_v2()
    - Sinon → parse_tableau_formateurs() (ancien format matriciel)

    Permet une transition sans rupture : l'ancien fichier continue de
    fonctionner tant qu'il n'est pas remplacé.
    """
    try:
        wb     = load_workbook(filepath, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        if 'AFFECTATIONS' in sheets:
            return parse_tableau_formateurs_v2(filepath)
    except Exception:
        pass
    return parse_tableau_formateurs(filepath)


# ─── Moteur assignation ───────────────────────────────────────────────────────
def assigner(planning_classes, dispos_formateurs, affectations):
    dispo_idx = {d['nom']: d['dispo'] for d in dispos_formateurs}
    result = defaultdict(dict)
    stats = {'assigned': 0, 'warn': 0}

    for ci in planning_classes:
        cn = ci['nom']; jours = ci['jours']
        aff = next((v for k, v in affectations.items()
                    if k.strip().lower() in cn.lower() or cn.lower() in k.strip().lower()), None)
        if not aff:
            for j in jours: result[j][cn] = {'formateur': '?', 'matiere': '?', 'slot': 'matin'}
            continue
        si = 0
        for j in jours:
            slot = ['matin', 'pm'][si % 2]; si += 1; assigned = None
            for e in sorted(aff, key=lambda x: x['heures_faites']):
                pd_ = dispo_idx.get(e['formateur'], {}); jd = pd_.get(j, {})
                if e['heures'] - e['heures_faites'] <= 0: continue
                if not pd_ or jd.get(slot, False): assigned = e; break
            if assigned:
                assigned['heures_faites'] += 4
                result[j][cn] = {'formateur': assigned['formateur'], 'matiere': assigned['matiere'], 'slot': slot}
                stats['assigned'] += 1
            else:
                result[j][cn] = {'formateur': '⚠️', 'matiere': '', 'slot': slot}
                stats['warn'] += 1

    heures = defaultdict(lambda: defaultdict(int))
    for dv in result.values():
        for cl, inf in dv.items():
            if inf['formateur'] not in ['?', '⚠️']:
                heures[inf['formateur']][inf['matiere']] += 4
    return dict(result), stats, dict(heures)

# ─── Écriture planning assigné ────────────────────────────────────────────────
def ecrire_planning(template_path, assignment, mois_cibles, output_path):
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    kw = ['BTS','BAC','EC ','CGC','NDRC','GPME','RDC','RH','Master']
    jf = {'lundi':0,'mardi':1,'mercredi':2,'jeudi':3,'vendredi':4,'lun':0,'mar':1,'mer':2,'jeu':3,'ven':4}

    for sn in wb.sheetnames:
        ws = wb[sn]; titre = ''
        # PATCH 8 — lecture titre robuste : cherche une cellule contenant un nom de mois
        # (pas juste la première cellule avec len>3, qui peut être le nom de l'école)
        for ri in range(1, 5):
            for ci in range(1, ws.max_column + 1):
                v = ws.cell(row=ri, column=ci).value
                if not (v and isinstance(v, str)): continue
                for m_key in MOIS_ABBR:
                    if m_key.upper() in v.upper():
                        titre = v.upper()
                        break
                if titre: break
            if titre: break
        # Fallback : utiliser le nom de l'onglet (ex: "Sept 2026")
        if not titre:
            titre = sn.upper()
        mn, yr = None, None
        for m, mnum in MOIS_ABBR.items():
            if m.upper() in titre:
                y = re.search(r'(20\d\d)', titre)
                if y: yr = int(y.group(1)); mn = mnum; break
        if not mn or f'{mn}/{yr}' not in mois_cibles: continue

        cc = {}
        for ri in range(1, 6):
            for ci in range(1, ws.max_column+1):
                v = ws.cell(row=ri, column=ci).value
                if isinstance(v, str) and any(k in v for k in kw): cc[v.strip()] = ci

        dr = {}
        for ri in range(1, ws.max_row+1):
            for ci in range(1, 4):
                v = ws.cell(row=ri, column=ci).value
                if isinstance(v, str) and v.strip().lower() in jf:
                    for lri in [ri, ri+1]:
                        for lci in range(1, 4):
                            dv = ws.cell(row=lri, column=lci).value
                            if isinstance(dv, (int, float)) and 1 <= int(dv) <= 31:
                                try:
                                    d = datetime.date(yr, mn, int(dv))
                                    ds = d.strftime('%Y-%m-%d')
                                    if ds not in dr: dr[ds] = ri
                                except: pass

        for ds, ri in dr.items():
            if ds not in assignment: continue
            for cn, ci in cc.items():
                for k, v in assignment[ds].items():
                    if k.strip().lower() in cn.lower() or cn.lower() in k.strip().lower():
                        cell = ws.cell(row=ri, column=ci)
                        cell.value = '⚠️' if v['formateur'] == '⚠️' else f"{v['formateur']}\n{v['matiere']}"
                        old = cell.font
                        cell.font = Font(name=old.name or 'Calibri', size=old.size or 9, bold=True, color=old.color)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        break
    wb.save(output_path)

# ─── Génération template colorié (ancien) ────────────────────────────────────
def detect_structure(ws):
    kw = ['BTS','BAC','EC ','CGC','NDRC','GPME','RDC','RH','Master']
    cc = {}; colors = {}
    for ri in range(1, 6):
        for ci in range(1, ws.max_column+1):
            cell = ws.cell(row=ri, column=ci); v = cell.value
            if isinstance(v, str) and any(k in v for k in kw):
                nom = v.strip(); cc[nom] = ci
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                    c = cell.fill.fgColor.rgb
                    if c not in ('00000000', 'FFFFFFFF'): colors[nom] = c
    for nom, ci in cc.items():
        if nom in colors: continue
        for ri in range(5, min(25, ws.max_row+1)):
            cell = ws.cell(row=ri, column=ci)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                c = cell.fill.fgColor.rgb
                if c not in ('00000000', 'FFFFFFFF', None):
                    colors[nom] = c; break
    jf_set = {'lundi','mardi','mercredi','jeudi','vendredi','lun','mar','mer','jeu','ven'}
    fdr = 6; dlc = 2
    for ri in range(3, 20):
        for ci in range(1, 5):
            v = ws.cell(row=ri, column=ci).value
            if isinstance(v, str) and v.strip().lower() in jf_set:
                fdr = ri; dlc = ci; break
        else: continue
        break
    return {'class_cols': cc, 'class_colors': colors, 'first_data_row': fdr, 'day_label_col': dlc}

def colorier_multifeuilles(multi_path, planning_classes, output_path):
    """
    Nouveau flux : reçoit un Excel multi-feuilles (produit par generer_excel_multifeuilles)
    et colorie directement les cases de chaque classe jour par jour.

    - Détecte le mois/année depuis le titre de chaque feuille (ex. "Septembre 2026")
    - Détecte la structure (colonnes classes, ligne de début, colonne jour) via detect_structure
    - Repère les lignes de données en cherchant le nom du jour + numéro dans les premières colonnes
    - Colorie les cases des classes qui ont cours ce jour-là
    """
    from openpyxl.worksheet.cell_range import CellRange as _CR
    shutil.copy(multi_path, output_path)
    wb = load_workbook(output_path)

    # Pré-indexer les jours de cours par classe : { 'YYYY-MM-DD' : set(noms_classes) }
    jours_par_date = defaultdict(set)
    for pc in planning_classes:
        for ds in pc.get('jours', []):
            jours_par_date[ds].add(pc['nom'])

    jf = {'lundi':0,'mardi':1,'mercredi':2,'jeudi':3,'vendredi':4,
          'lun':0,'mar':1,'mer':2,'jeu':3,'ven':4}
    total = 0

    for sn in wb.sheetnames:
        ws = wb[sn]

        # Détecter mois/année depuis le titre de la feuille ou cellules en-tête
        mn, yr = None, None
        for src in [sn] + [ws.cell(row=r, column=c).value
                            for r in range(1, 4) for c in range(1, ws.max_column+1)
                            if ws.cell(row=r, column=c).value]:
            if not isinstance(src, str): continue
            for m_key, m_num in MOIS_ABBR.items():
                if m_key.upper() in src.upper():
                    y = re.search(r'(20\d\d)', src)
                    if y:
                        mn = m_num; yr = int(y.group(1)); break
            if mn: break
        if not mn:
            continue

        struct = detect_structure(ws)
        cc = struct['class_cols']; colors = struct['class_colors']
        ci_def = 0
        for nom in cc:
            if nom not in colors:
                colors[nom] = DEFAULT_COLORS[ci_def % len(DEFAULT_COLORS)]; ci_def += 1

        # Repérer les lignes de données : cherche nom de jour dans col 1-4
        row_dates = {}  # { 'YYYY-MM-DD': row_idx }
        for ri in range(1, ws.max_row + 1):
            for ci2 in range(1, 5):
                v = ws.cell(row=ri, column=ci2).value
                if not isinstance(v, str): continue
                if v.strip().lower() not in jf: continue
                # Cherche le numéro du jour dans les colonnes proches
                for lci in range(1, 6):
                    dv = ws.cell(row=ri, column=lci).value
                    if isinstance(dv, (int, float)) and 1 <= int(dv) <= 31:
                        try:
                            d = datetime.date(yr, mn, int(dv))
                            ds = d.strftime('%Y-%m-%d')
                            if ds not in row_dates:
                                row_dates[ds] = ri
                        except: pass
                        break
                break

        # Colorier
        for ds, ri in row_dates.items():
            classes_ce_jour = jours_par_date.get(ds, set())
            for classe_nom, col_idx in cc.items():
                a_cours = any(
                    pc_nom.strip().lower() in classe_nom.lower()
                    or classe_nom.lower() in pc_nom.strip().lower()
                    for pc_nom in classes_ce_jour
                )
                cell = ws.cell(row=ri, column=col_idx)
                if a_cours:
                    color = colors.get(classe_nom, 'FFD9D9D9')
                    if len(color) == 6: color = 'FF' + color
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    total += 1
                else:
                    cell.fill = PatternFill(fill_type=None)
                    cell.value = None

    wb.save(output_path)
    return total, len(wb.sheetnames)


def generer_template_colorie(template_path, planning_classes, annee_debut, output_path, mois_cibles=None):
    wb_tpl = load_workbook(template_path)
    struct = detect_structure(wb_tpl.active)
    cc = struct['class_cols']; colors = struct['class_colors']
    fdr = struct['first_data_row']; dlc = struct['day_label_col']

    ci = 0
    for nom in cc:
        if nom not in colors:
            colors[nom] = DEFAULT_COLORS[ci % len(DEFAULT_COLORS)]; ci += 1

    mois_scolaire_all = [(9,annee_debut),(10,annee_debut),(11,annee_debut),(12,annee_debut),
                          (1,annee_debut+1),(2,annee_debut+1),(3,annee_debut+1),(4,annee_debut+1),
                          (5,annee_debut+1),(6,annee_debut+1),(7,annee_debut+1),(8,annee_debut+1)]
    if mois_cibles:
        cibles_set = {(int(m2), int(a2)) for a2, m2 in mois_cibles}
        mois_scolaire = [(m2, a2) for m2, a2 in mois_scolaire_all if (m2, a2) in cibles_set]
    else:
        mois_scolaire = mois_scolaire_all
    jf_nom = {0:'Lundi',1:'Mardi',2:'Mercredi',3:'Jeudi',4:'Vendredi'}

    wb_out = Workbook(); wb_out.remove(wb_out.active)
    total = 0

    for (mois, annee) in mois_scolaire:
        wb_tmp = load_workbook(template_path); ws_src = wb_tmp.active
        ws = wb_out.create_sheet(title=f"{MOIS_FR[mois]} {annee}")

        for row in ws_src.iter_rows():
            for cell in row:
                nc = ws.cell(row=cell.row, column=cell.column)
                nc.value = cell.value
                if cell.has_style:
                    nc.font = copy.copy(cell.font); nc.fill = copy.copy(cell.fill)
                    nc.border = copy.copy(cell.border); nc.alignment = copy.copy(cell.alignment)
                    nc.number_format = cell.number_format
        for cl, cd in ws_src.column_dimensions.items():
            ws.column_dimensions[cl].width = cd.width
        for ri, rd in ws_src.row_dimensions.items():
            ws.row_dimensions[ri].height = rd.height
        # APRÈS — Bug #2 corrigé : bypass ws.merge_cells() O(n²) → CellRange direct
        from openpyxl.worksheet.cell_range import CellRange as _CR
        _seen = set()
        for mg in ws_src.merged_cells.ranges:
            k = str(mg)
            if k not in _seen:
                _seen.add(k)
                try: ws.merged_cells.ranges.add(_CR(k))
                except: pass

        for ri in range(1, 4):
            for ci2 in range(1, ws.max_column+1):
                v = ws.cell(row=ri, column=ci2).value
                if isinstance(v, str) and re.search(r'20\d\d', v):
                    ws.cell(row=ri, column=ci2).value = f"{MOIS_FR[mois].upper()} {annee}"

        nb_j = calendar.monthrange(annee, mois)[1]
        jours_ouvres = [datetime.date(annee, mois, j) for j in range(1, nb_j+1)
                        if datetime.date(annee, mois, j).weekday() < 5
                        and datetime.date(annee, mois, j) not in FERIES]

        row_idx = fdr
        for d in jours_ouvres:
            ds = d.strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=dlc).value = jf_nom[d.weekday()]
            ws.cell(row=row_idx, column=dlc+1).value = d.day
            for classe_nom, col_idx in cc.items():
                a_cours = any(
                    ds in pc['jours']
                    for pc in planning_classes
                    if pc['nom'] and (pc['nom'].strip().lower() in classe_nom.lower()
                                      or classe_nom.lower() in pc['nom'].strip().lower())
                )
                cell = ws.cell(row=row_idx, column=col_idx)
                if a_cours:
                    color = colors.get(classe_nom, 'FFD9D9D9')
                    if len(color) == 6: color = 'FF' + color
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    total += 1
                else:
                    cell.fill = PatternFill(fill_type=None); cell.value = None
            row_idx += 1

    wb_out.save(output_path)
    return total, len(mois_scolaire)

# ─── Stratégies de génération d'un mois ──────────────────────────────────────
#
# Deux stratégies disponibles, sélectionnables via le paramètre `mode` :
#   "delete" (défaut) : supprime physiquement les lignes inutiles — rendu propre,
#                       mais lent sur les templates riches en merged cells (O(n²))
#   "hide"            : masque les lignes inutilisées avec height=0 — instantané,
#                       structure fixe préservée, fichier légèrement plus lourd
#
# Interface commune :
#   generate_month_sheet(ws, annee, mois, mode) → int (nb jours ouvrés)
#   generer_template_mois(template_path, output_path, annee, mois, mode) → int
#   generer_excel_multifeuilles(template_path, mois_liste, output_path, mode) → int

def _ecrire_titre_mois(ws, annee, mois):
    """Remplace le mois/année dans les cellules titre (lignes 1-3)."""
    import re as _re
    nom_mois = MOIS_FR_UPPER[mois]
    pattern  = (r'(?:JANVIER|FÉVRIER|FEVRIER|MARS|AVRIL|MAI|JUIN|JUILLET|'
                r'AOÛT|AOUT|SEPTEMBRE|OCTOBRE|NOVEMBRE|DÉCEMBRE|DECEMBRE) 20\d\d')
    for r in range(1, 4):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=col).value
            if not isinstance(v, str): continue
            if _re.search(pattern, v.upper()):
                ws.cell(row=r, column=col).value = _re.sub(
                    pattern, f"{nom_mois} {annee}", v.upper(), flags=_re.IGNORECASE)

def _jours_ouvres_mois(annee, mois):
    """Retourne la liste des (label_jour, num_jour) ouvrés du mois."""
    jours = []
    for j in range(1, calendar.monthrange(annee, mois)[1] + 1):
        d = datetime.date(annee, mois, j)
        if d.weekday() < 5:
            jours.append((JOURS_FR_LIST[d.weekday()], j))
    return jours

def _slot_debut(annee, mois):
    """Nombre de slots à sauter en début de mois (0 = lundi, 4 = vendredi)."""
    premier = datetime.date(annee, mois, 1)
    return premier.weekday() if premier.weekday() < 5 else 0

def _detecter_slots(ws):
    """Détecte les lignes label de chaque slot jour dans la feuille."""
    slots = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and any(j in v for j in JOURS_FR_LIST):
            slots.append(r)
    return slots


def generate_month_sheet_delete(ws, annee, mois):
    """
    Stratégie DELETE : supprime physiquement les lignes de slots inutilisés.
    Rendu Excel propre (pas de lignes cachées).
    Coût : delete_rows openpyxl est O(n²) sur les merged cells → lent sur gros templates.
    """
    import re as _re
    from openpyxl.worksheet.cell_range import CellRange

    jours_ouvres = _jours_ouvres_mois(annee, mois)
    if not jours_ouvres:
        raise ValueError(f"Aucun jour ouvré pour {mois}/{annee}")
    slot_deb  = _slot_debut(annee, mois)
    lignes_debut = slot_deb * 6

    _ecrire_titre_mois(ws, annee, mois)

    # ── Supprimer les slots vides au DÉBUT ────────────────────────────────────
    if lignes_debut > 0:
        DELETE_START = 7
        DELETE_COUNT = lignes_debut
        FIRST_KEPT   = DELETE_START + DELETE_COUNT
        saved_heights = {r: ws.row_dimensions[r].height
                         for r in range(FIRST_KEPT, ws.max_row + 1)}
        ws.delete_rows(DELETE_START, DELETE_COUNT)
        for old_r, h in saved_heights.items():
            new_r = old_r - DELETE_COUNT
            if new_r >= DELETE_START and h is not None:
                ws.row_dimensions[new_r].height = h
        target_rows = {r for r in range(DELETE_START, ws.max_row + 1)
                       if (ws.row_dimensions[r].height is None
                           or ws.row_dimensions[r].height >= 10)}
        for mg in list(ws.merged_cells.ranges):
            if mg.min_row == mg.max_row and mg.min_row in target_rows:
                try: ws.merged_cells.ranges.discard(mg)
                except Exception: pass
        for r in target_rows:
            for c1, c2 in ALL_MERGE_PAIRS:
                ws.merged_cells.ranges.add(
                    CellRange(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}"))

    # ── Écrire les labels dans les slots restants ─────────────────────────────
    jours_pos = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and any(j in v for j in JOURS_FR_LIST):
            jours_pos.append((r, r + 1))

    for i, (rl, rn) in enumerate(jours_pos):
        if i < len(jours_ouvres):
            label, num = jours_ouvres[i]
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = label
                ws.cell(row=rn, column=col).value = num
        else:
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = None
                ws.cell(row=rn, column=col).value = None

    # ── Supprimer les slots vides à la FIN ────────────────────────────────────
    grey_closing_row = None
    for r in range(ws.max_row, 0, -1):
        fill = ws.cell(row=r, column=2).fill
        bg = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == 'rgb' else None
        if bg == 'FFC0C0C0':
            grey_closing_row = r; break

    if len(jours_pos) > len(jours_ouvres):
        last_label_row = jours_pos[len(jours_ouvres) - 1][0]
        delete_from    = last_label_row + 5
        delete_until   = (grey_closing_row - 1) if grey_closing_row else ws.max_row
        if delete_from <= delete_until:
            ws.delete_rows(delete_from, delete_until - delete_from + 1)

    return len(jours_ouvres)


def generate_month_sheet_hide(ws, annee, mois):
    """
    Stratégie HIDE : masque les lignes de slots inutilisés avec height=0.
    Aucun delete_rows → rapide même sur les templates riches en merged cells.
    Les lignes masquées sont invisibles dans Excel mais présentes dans le fichier.
    """
    jours_ouvres = _jours_ouvres_mois(annee, mois)
    if not jours_ouvres:
        raise ValueError(f"Aucun jour ouvré pour {mois}/{annee}")
    slot_deb = _slot_debut(annee, mois)

    _ecrire_titre_mois(ws, annee, mois)

    slots = _detecter_slots(ws)

    for i, rl in enumerate(slots):
        jour_idx = i - slot_deb
        if 0 <= jour_idx < len(jours_ouvres):
            label, num = jours_ouvres[jour_idx]
            for col in JOURS_COLS:
                ws.cell(row=rl,     column=col).value = label
                ws.cell(row=rl + 1, column=col).value = num
            # S'assurer que le slot est visible (reset height=0 éventuel)
            for offset in range(6):
                rd = ws.row_dimensions[rl + offset]
                if rd.height is not None and rd.height < 1:
                    rd.height = None
        else:
            # Slot inutilisé : effacer les valeurs et marquer pour suppression XML
            for offset in range(6):
                ws.cell(row=rl + offset, column=1).value = '__DEL__'
            for col in JOURS_COLS:
                ws.cell(row=rl,     column=col).value = None
                ws.cell(row=rl + 1, column=col).value = None

    return len(jours_ouvres)


# Dispatcher : sélectionne la stratégie selon `mode`
_STRATEGIES = {
    'delete': generate_month_sheet_delete,
    'hide':   generate_month_sheet_hide,
}

def generate_month_sheet(ws, annee, mois, mode='delete'):
    """
    Point d'entrée unique pour la génération d'un mois sur une feuille openpyxl.
    mode : 'delete' (défaut) ou 'hide'
    """
    strategy = _STRATEGIES.get(mode, generate_month_sheet_delete)
    return strategy(ws, annee, mois)


# ─── Génération template vierge par mois ─────────────────────────────────────
def _supprimer_lignes_masquees_xml(xlsx_path):
    """
    Post-traitement XML : supprime physiquement les lignes marquées '__DEL__' en colonne A.
    Bypasse openpyxl (delete_rows O(n²)) en manipulant directement le ZIP/XML.
    Les merged cells qui chevauchent ces lignes sont aussi retirées.
    """
    import zipfile, io
    from xml.etree import ElementTree as ET

    NS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    ET.register_namespace('', NS)

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        names = zin.namelist()
        blobs = {n: zin.read(n) for n in names}

    # 1. Trouver l'index de '__DEL__' dans les shared strings
    sst_name = next((n for n in names if n.endswith('sharedStrings.xml')), None)
    del_idx  = None
    if sst_name:
        sst = ET.fromstring(blobs[sst_name])
        for i, si in enumerate(sst):
            t = si.find(f'{{{NS}}}t')
            if t is not None and t.text == '__DEL__':
                del_idx = str(i)
                break

    if del_idx is None:
        return  # pas de marqueur → rien à faire

    # 2. Traiter chaque feuille
    sheet_names = sorted(
        [n for n in names if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')],
        key=lambda x: int(re.search(r'(\d+)', x.split('/')[-1]).group(1))
    )

    rows_tag   = f'{{{NS}}}row'
    cell_tag   = f'{{{NS}}}c'
    merged_tag = f'{{{NS}}}mergeCells'
    mc_tag     = f'{{{NS}}}mergeCell'

    for sheet_name in sheet_names:
        tree      = ET.fromstring(blobs[sheet_name])
        sheetData = tree.find(f'{{{NS}}}sheetData')
        if sheetData is None:
            continue

        hidden_rows    = set()
        rows_to_remove = []

        for row_el in sheetData:
            if row_el.tag != rows_tag:
                continue
            # Cherche une cellule en colonne A (r="A<n>") avec valeur = del_idx
            for cell_el in row_el:
                if cell_el.tag != cell_tag:
                    continue
                coord = cell_el.get('r', '')
                if not re.match(r'^A\d+$', coord):
                    continue
                if cell_el.get('t') == 's' and cell_el.find(f'{{{NS}}}v') is not None:
                    if cell_el.find(f'{{{NS}}}v').text == del_idx:
                        hidden_rows.add(int(row_el.get('r', 0)))
                        rows_to_remove.append(row_el)
                break

        for row_el in rows_to_remove:
            sheetData.remove(row_el)

        if not hidden_rows:
            continue  # feuille inchangée

        # 3. Renuméroter les lignes restantes
        sorted_hidden = sorted(hidden_rows)

        def _shift(r):
            return r - sum(1 for h in sorted_hidden if h < r)

        for row_el in sheetData:
            if row_el.tag != rows_tag:
                continue
            old_r = int(row_el.get('r'))
            new_r = _shift(old_r)
            if new_r != old_r:
                row_el.set('r', str(new_r))
                for cell_el in row_el:
                    coord = cell_el.get('r', '')
                    col_part = ''.join(c for c in coord if c.isalpha())
                    if col_part:
                        cell_el.set('r', f'{col_part}{new_r}')

        # 4. Nettoyer les mergeCells
        for mc_parent in tree.iter(merged_tag):
            to_remove = []
            for mc in mc_parent:
                if mc.tag != mc_tag:
                    continue
                ref = mc.get('ref', '')
                m2 = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', ref)
                if not m2:
                    continue
                c1, r1, c2, r2 = m2.group(1), int(m2.group(2)), m2.group(3), int(m2.group(4))
                if any(r1 <= h <= r2 for h in hidden_rows):
                    to_remove.append(mc)
                else:
                    nr1, nr2 = _shift(r1), _shift(r2)
                    if nr1 != r1 or nr2 != r2:
                        mc.set('ref', f'{c1}{nr1}:{c2}{nr2}')
            for mc in to_remove:
                mc_parent.remove(mc)

        blobs[sheet_name] = ET.tostring(tree, encoding='UTF-8', xml_declaration=True)

    # 5. Réécrire le ZIP avec toutes les feuilles mises à jour
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, blobs[n])
    buf.seek(0)
    with open(xlsx_path, 'wb') as f:
        f.write(buf.read())


def generer_template_mois(template_path, output_path, annee, mois, mode='delete'):
    """
    Génère un template vierge pour un mois donné dans output_path.
    mode : 'delete' (défaut) ou 'hide'
    En mode delete : génère d'abord en hide (rapide), puis supprime les lignes
    masquées via XML direct (évite delete_rows O(n²) d'openpyxl).
    """
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    effective_mode = 'hide' if mode == 'delete' else mode
    nb = generate_month_sheet(ws, annee, mois, mode=effective_mode)
    wb.save(output_path)
    if mode == 'delete':
        _supprimer_lignes_masquees_xml(output_path)
    return nb


# ─── Excel multi-feuilles (une feuille par mois) ─────────────────────────────
def _copier_feuille(ws_src, ws_dst):
    """
    Copie complète d'une feuille vers une autre (workbooks différents).
    Recrée les styles via leurs constructeurs pour éviter copy.copy() cassé sur Python 3.14+.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    for row in ws_src.iter_rows():
        for cell in row:
            nc = ws_dst.cell(row=cell.row, column=cell.column)
            nc.value = cell.value
            if cell.has_style:
                f = cell.font
                nc.font = Font(
                    name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                    underline=f.underline, color=copy.copy(f.color),
                    strike=f.strike, vertAlign=f.vertAlign
                )
                fi = cell.fill
                nc.fill = PatternFill(
                    fill_type=fi.fill_type,
                    fgColor=copy.copy(fi.fgColor),
                    bgColor=copy.copy(fi.bgColor)
                )
                b = cell.border
                def _side(s):
                    return Side(border_style=s.border_style, color=copy.copy(s.color))
                nc.border = Border(
                    left=_side(b.left), right=_side(b.right),
                    top=_side(b.top), bottom=_side(b.bottom)
                )
                a = cell.alignment
                nc.alignment = Alignment(
                    horizontal=a.horizontal, vertical=a.vertical,
                    wrap_text=a.wrap_text, indent=a.indent
                )
                nc.number_format = cell.number_format
    for col, cd in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col].width = cd.width
    for r, rd in ws_src.row_dimensions.items():
        if rd.height:
            ws_dst.row_dimensions[r].height = rd.height
    # Bypass merge_cells() qui fait une vérification __contains__ O(n²) très lente
    # On ajoute les fusions directement dans la collection interne
    from openpyxl.worksheet.cell_range import CellRange
    seen = set()
    for mg in ws_src.merged_cells.ranges:
        k = str(mg)
        if k not in seen:
            seen.add(k)
            try:
                ws_dst.merged_cells.ranges.add(CellRange(k))
            except Exception:
                pass


def generer_excel_multifeuilles(template_path, mois_liste, output_path, mode='delete'):
    """
    Génère un Excel multi-feuilles (1 onglet par mois).
    mode : 'delete' (défaut) ou 'hide'
    Stratégie : génère chaque mois dans un fichier temporaire puis fusionne.
    """
    import tempfile

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    total_jours = 0

    for (annee, mois) in mois_liste:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        nb_jours = generer_template_mois(template_path, tmp_path, annee, mois, mode=mode)
        total_jours += nb_jours

        wb_tmp = load_workbook(tmp_path)
        ws_src = wb_tmp.active
        sheet_name = f"{MOIS_FR[mois][:4]} {annee}"
        ws_dst = wb_out.create_sheet(title=sheet_name)
        if mode == 'hide':
            _copier_feuille_rapide(ws_src, ws_dst)
        else:
            _copier_feuille(ws_src, ws_dst)
        wb_tmp.close()
        os.unlink(tmp_path)

    wb_out.save(output_path)
    return total_jours

# ─── Fusion de fichiers Excel ─────────────────────────────────────────────────
def _copier_feuille_rapide(ws_src, ws_dst):
    """
    Copie rapide pour la fusion — sans copy.copy() (instable selon version openpyxl).
    Stratégie par priorité décroissante :
      - cellule avec valeur  → styles complets (font + fill + alignment + number_format)
      - cellule vide colorée → fill seul (préserve le code couleur des jours de cours)
      - cellule vide non colorée → ignorée (bordures seules, sans intérêt pour la fusion)
    """
    from openpyxl.worksheet.cell_range import CellRange
    from openpyxl.styles import Font, PatternFill, Alignment
    MAX_COL = 130  # colonne DZ
    NO_FILL = {'none', None, ''}

    for (row, col), cell in ws_src._cells.items():
        if col > MAX_COL:
            continue
        has_value = cell.value is not None
        fi = cell.fill if cell.has_style else None
        has_color = fi and fi.fill_type not in NO_FILL

        if not has_value and not has_color:
            continue  # bordure seule : on saute

        nc = ws_dst.cell(row=row, column=col, value=cell.value)

        if has_color or (has_value and fi):
            nc.fill = PatternFill(fill_type=fi.fill_type,
                                  fgColor=copy.copy(fi.fgColor),
                                  bgColor=copy.copy(fi.bgColor))
        if has_value and cell.has_style:
            f = cell.font
            nc.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                           underline=f.underline, color=copy.copy(f.color),
                           strike=f.strike, vertAlign=f.vertAlign)
            a = cell.alignment
            nc.alignment = Alignment(horizontal=a.horizontal, vertical=a.vertical,
                                     wrap_text=a.wrap_text, indent=a.indent)
            nc.number_format = cell.number_format

    for col, cd in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col].width = cd.width
    for r, rd in ws_src.row_dimensions.items():
        if rd.height is not None:
            ws_dst.row_dimensions[r].height = rd.height
    seen = set()
    for mg in ws_src.merged_cells.ranges:
        k = str(mg)
        if k not in seen:
            seen.add(k)
            try:
                ws_dst.merged_cells.ranges.add(CellRange(k))
            except Exception:
                pass


def fusionner_excels(sources, noms_feuilles=None):
    """
    Fusionne plusieurs fichiers Excel (une feuille chacun) en un seul fichier multi-feuilles.
    sources       : list[bytes | str | Path]
    noms_feuilles : list[str] optionnel — noms des onglets dans le fichier final
    Retourne bytes.
    """
    from io import BytesIO
    wb_dest = Workbook()
    wb_dest.remove(wb_dest.active)
    used_names = []

    for i, source in enumerate(sources):
        if isinstance(source, (bytes, bytearray)):
            wb_src = load_workbook(BytesIO(source))
        else:
            wb_src = load_workbook(source)

        ws_src = wb_src.active
        raw = (noms_feuilles[i].strip()
               if noms_feuilles and i < len(noms_feuilles) and noms_feuilles[i].strip()
               else ws_src.title)
        nom = raw[:31]
        if nom in used_names:
            nom = f"{nom[:27]}_{i+1}"
        used_names.append(nom)

        ws_dst = wb_dest.create_sheet(title=nom)
        _copier_feuille_rapide(ws_src, ws_dst)
        wb_src.close()

    out = BytesIO()
    wb_dest.save(out)
    return out.getvalue()

# ─── Routes ───────────────────────────────────────────────────────────────────
@app.route('/ping')
def ping():
    return jsonify({'ok': True})

@app.route('/generer', methods=['POST'])
def generer():
    sid = str(uuid.uuid4())[:8]; wd = os.path.join(UPLOAD_FOLDER, sid); os.makedirs(wd, exist_ok=True)
    def save(f, sub=''):
        d = os.path.join(wd, sub) if sub else wd; os.makedirs(d, exist_ok=True)
        p = os.path.join(d, f.filename); f.save(p); return p
    try:
        # Compat nouveau front (planning_0/N, disponibilites, formateurs)
        # vs ancien front (classes getlist, dispos getlist, formateurs)
        cf_new = [request.files[k] for k in sorted(request.files) if k.startswith('planning_')]
        cf_old = request.files.getlist('classes')
        cf = cf_new if cf_new else cf_old

        # disponibilites : nouveau front envoie 1 fichier 'disponibilites'
        # ancien front envoyait une liste 'dispos'
        dispos_file = request.files.get('disponibilites')
        df_old = request.files.getlist('dispos')

        ff = request.files.get('formateurs'); tf = request.files.get('template')
        if not all([cf, ff, tf]) or (not dispos_file and not df_old):
            return jsonify({'error': 'Fichiers manquants (template, disponibilites, formateurs, plannings requis)'}), 400

        # Plage de mois : nouveau front (annee_debut/mois_debut/annee_fin/mois_fin)
        annee_debut = int(request.form.get('annee_debut', 2026))
        mois_debut  = int(request.form.get('mois_debut',  9))
        annee_fin   = int(request.form.get('annee_fin',   annee_debut + 1))
        mois_fin    = int(request.form.get('mois_fin',    8))

        # APRÈS — Bug #3 corrigé : priorité à mois_json (liste exacte)
        # Bug #4 latent corrigé : mois_old avait annee/mois inversés (sans impact
        #   avec le nouveau front qui n'envoie jamais 'mois')
        mois_json = request.form.get('mois_json')
        if mois_json:
            mois_liste = [(int(a), int(m)) for a, m in json.loads(mois_json)][:12]
        else:
            # Fallback plage (ancien front ou appel direct)
            mois_liste = []
            a, m = annee_debut, mois_debut
            while (a, m) <= (annee_fin, mois_fin):
                mois_liste.append((a, m))
                m += 1
                if m > 12: m = 1; a += 1
                if len(mois_liste) > 12: break
            # Ancien format mois set JSON "mois/annee" (compatibilité)
            mois_old = set(json.loads(request.form.get('mois', '[]')))
            if mois_old:
                # Format correct : "9/2026" → (annee=2026, mois=9)
                mois_liste = [(int(mo.split('/')[1]), int(mo.split('/')[0])) for mo in mois_old]

        if not mois_liste: return jsonify({'error': 'Plage de mois invalide'}), 400

        cp = [save(f, 'classes') for f in cf if f.filename]
        fp = save(ff); tp = save(tf)

        if dispos_file:
            dp = [save(dispos_file, 'dispos')]
        else:
            dp = [save(f, 'dispos') for f in df_old if f.filename]

        pcs    = [cl for cl in [parse_planning_classe(p) for p in cp] if cl]
        dispos = [parse_disponibilite(p) for p in dp]
        aff    = _auto_parse_formateurs(fp)
        # Extraire le log debug avant de passer aff à assigner()
        # (_debug n'est présent qu'avec le nouveau parseur v2)
        aff_debug = aff.pop('_debug', {})
        assignment, stats, heures = assigner(pcs, dispos, aff)

        mois_str = ', '.join(f"{MOIS_FR[m][:3]} {a}" for a, m in mois_liste)
        mois_set = {f"{m}/{a}" for a, m in mois_liste}
        on = f"Planning_Formateurs_{annee_debut}_{annee_fin}.xlsx"
        op = os.path.join(wd, on)

        # APRÈS — Bug #1 corrigé : ecrire_planning attend un fichier multi-feuilles
        # (une feuille par mois avec le nom du mois dans le titre de cellule).
        # On génère d'abord le template vierge multi-feuilles, puis on y écrit les assignations.
        tmp_multi = os.path.join(wd, '_tmp_multi.xlsx')
        generer_excel_multifeuilles(tp, mois_liste, tmp_multi)
        ecrire_planning(tmp_multi, assignment, mois_set, op)
        try: os.unlink(tmp_multi)
        except: pass
        # Info de debug exposées dans le JSON pour diagnostiquer les feuilles vides
        classes_noms   = [p['nom'] for p in pcs if p.get('nom')]
        jours_trouves  = sorted(list(assignment.keys()))[:5]  # 5 premiers jours
        return jsonify({
            'sessions_assignees': stats['assigned'], 'creneaux_sans_prof': stats['warn'],
            'formateurs_actifs': len(heures), 'classes': len(pcs),
            'mois': mois_str, 'nb_mois': len(mois_liste),
            'fichier': on, 'session_id': sid,
            'heures_formateurs': {p: dict(m) for p, m in heures.items()},
            'debug_classes': classes_noms,
            'debug_jours_assignes': jours_trouves,
            'debug_nb_jours': len(assignment),
            'debug_affectations': aff_debug,   # classes/formateurs/affectations détectés
        })
    except Exception as e:
        import traceback; return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/generer-template', methods=['POST'])
def generer_template_colorie_route():
    sid = str(uuid.uuid4())[:8]; wd = os.path.join(UPLOAD_FOLDER, sid); os.makedirs(wd, exist_ok=True)
    def save(f, sub=''):
        d = os.path.join(wd, sub) if sub else wd; os.makedirs(d, exist_ok=True)
        p = os.path.join(d, f.filename); f.save(p); return p
    try:
        # Compat nouveau front : planning_0, planning_1, ...
        # Ancien front envoyait 'classes' (getlist) — on accepte les deux
        cf_new = [request.files[k] for k in sorted(request.files) if k.startswith('planning_')]
        cf_old = request.files.getlist('classes')
        cf = cf_new if cf_new else cf_old
        tf = request.files.get('template')
        if not cf or not tf: return jsonify({'error': 'Fichiers manquants'}), 400

        # Plage de mois : nouveau front (annee_debut/mois_debut/annee_fin/mois_fin)
        # ou ancien front (annee + mois_liste JSON)
        annee_debut = int(request.form.get('annee_debut', request.form.get('annee', 2026)))
        mois_debut  = int(request.form.get('mois_debut', 9))
        annee_fin   = int(request.form.get('annee_fin', annee_debut + (0 if mois_debut <= 8 else 0)))
        mois_fin    = int(request.form.get('mois_fin', 8))
        annee       = annee_debut  # compatibilité

        # APRÈS — Bug #3 corrigé : priorité à mois_json (liste exacte)
        mois_json = request.form.get('mois_json')
        mois_liste_json = request.form.get('mois_liste')
        if mois_json:
            mois_cibles = [(int(a2), int(m2)) for a2, m2 in json.loads(mois_json)][:12]
        elif mois_liste_json:
            mois_cibles = [(int(a2), int(m2)) for a2, m2 in json.loads(mois_liste_json)][:12]
        else:
            mois_cibles = []
            a, m = annee_debut, mois_debut
            while (a, m) <= (annee_fin, mois_fin):
                mois_cibles.append((a, m))
                m += 1
                if m > 12: m = 1; a += 1
                if len(mois_cibles) > 12: break
        if not mois_cibles: mois_cibles = None

        cp = [save(f, 'classes') for f in cf if f.filename]; tp = save(tf)
        pcs = [cl for cl in [parse_planning_classe(p) for p in cp] if cl]
        on = f"Template_Affichage_{annee}_{annee+1}.xlsx"; op = os.path.join(wd, on)

        # Détection automatique : multi-feuilles → nouveau flux, feuille unique → ancien flux
        wb_check = load_workbook(tp, read_only=True)
        is_multi = len(wb_check.sheetnames) > 1
        wb_check.close()

        if is_multi:
            total, nb = colorier_multifeuilles(tp, pcs, op)
        else:
            total, nb = generer_template_colorie(tp, pcs, annee, op, mois_cibles=mois_cibles)

        return jsonify({
            'fichier': on, 'session_id': sid, 'nb_mois': nb,
            'cases_colories': total, 'classes': len(pcs),
            'noms_classes': [cl['nom'] for cl in pcs if cl.get('nom')],
            'format': 'excel',
        })
    except Exception as e:
        import traceback; return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


def _appliquer_mois_sur_feuille(ws, annee, mois, grey_row_override=None):
    """
    Applique les transformations d'un mois sur une feuille déjà copiée du template.
    Réutilise la même logique que generer_template_mois mais sur un ws existant.
    """
    # Titre
    nom_mois = MOIS_FR_UPPER[mois]
    for r in range(1, 4):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and 'SEPTEMBRE' in v.upper():
                ws.cell(row=r, column=c).value = v.replace(
                    "SEPTEMBRE 2026", f"{nom_mois} {annee}")

    # Jours ouvrés
    jours_ouvres = []
    for j in range(1, calendar.monthrange(annee, mois)[1] + 1):
        d = datetime.date(annee, mois, j)
        if d.weekday() < 5 and d not in FERIES:
            jours_ouvres.append((JOURS_FR_LIST[d.weekday()], j))

    # Suppression lignes début
    premier = datetime.date(annee, mois, 1)
    while premier.weekday() >= 5 or premier in FERIES:
        premier += datetime.timedelta(days=1)
    lignes_a_supprimer = premier.weekday() * 6

    if lignes_a_supprimer > 0:
        DELETE_START = 7
        DELETE_COUNT = lignes_a_supprimer
        FIRST_KEPT   = DELETE_START + DELETE_COUNT
        saved_heights = {r: ws.row_dimensions[r].height
                         for r in range(FIRST_KEPT, ws.max_row + 1)}
        ws.delete_rows(DELETE_START, DELETE_COUNT)
        for old_r, h in saved_heights.items():
            new_r = old_r - DELETE_COUNT
            if new_r >= DELETE_START and h is not None:
                ws.row_dimensions[new_r].height = h
        # Fix fusions optimisé
        from openpyxl.worksheet.cell_range import CellRange
        target_rows = {r for r in range(DELETE_START, ws.max_row + 1)
                       if (ws.row_dimensions[r].height is None or ws.row_dimensions[r].height >= 10)}
        for mg in list(ws.merged_cells.ranges):
            if mg.min_row == mg.max_row and mg.min_row in target_rows:
                try: ws.merged_cells.ranges.discard(mg)
                except Exception: pass
        for r in target_rows:
            for c1, c2 in ALL_MERGE_PAIRS:
                ws.merged_cells.ranges.add(
                    CellRange(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}"))

    # Labels jours
    jours_pos = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and any(j in v for j in JOURS_FR_LIST):
            jours_pos.append((r, r + 1))

    for i, (rl, rn) in enumerate(jours_pos):
        if i < len(jours_ouvres):
            label, num = jours_ouvres[i]
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = label
                ws.cell(row=rn, column=col).value = num
        else:
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = None
                ws.cell(row=rn, column=col).value = None

    # Supprimer slots vides en fin (en préservant la ligne grise)
    # grey_row_override : position dans le template original (avant suppression lignes)
    # On la recalcule en tenant compte des lignes supprimées au début
    if grey_row_override is not None:
        grey_row = grey_row_override - lignes_a_supprimer
        if grey_row <= 0:
            grey_row = None
    else:
        grey_row = None
        for r in range(ws.max_row, 0, -1):
            try:
                fill = ws.cell(row=r, column=2).fill
                bg = (fill.fgColor.rgb
                      if fill and fill.fgColor and fill.fgColor.type == 'rgb'
                      else None)
                if bg == 'FFC0C0C0':
                    grey_row = r
                    break
            except Exception:
                pass
    if len(jours_pos) > len(jours_ouvres):
        last_rl = jours_pos[len(jours_ouvres) - 1][0]
        delete_from  = last_rl + 5
        delete_until = (grey_row - 1) if grey_row else ws.max_row
        if delete_from <= delete_until:
            ws.delete_rows(delete_from, delete_until - delete_from + 1)

    return len(jours_ouvres)


# ─── Système de jobs asynchrones ─────────────────────────────────────────────
import threading

JOBS      = {}
JOBS_LOCK = threading.Lock()

def _job_set(job_id, **kw):
    with JOBS_LOCK:
        JOBS[job_id].update(kw)

def _run_generer_template_vierge(job_id, tp, mois_liste, wd, sid, format_sortie, generation_mode):
    import zipfile as _zipfile, tempfile as _tf
    try:
        nb_mois     = len(mois_liste)
        total_jours = 0
        a_deb, m_deb = mois_liste[0]
        a_fin, m_fin = mois_liste[-1]
        label = (f"{MOIS_FR_UPPER[m_deb]} {a_deb}" if nb_mois == 1
                 else f"{MOIS_FR_UPPER[m_deb]} {a_deb} → {MOIS_FR_UPPER[m_fin]} {a_fin}")

        if format_sortie == 'excel' and nb_mois > 1:
            wb_out = Workbook()
            wb_out.remove(wb_out.active)
            for i, (annee, mois) in enumerate(mois_liste):
                _job_set(job_id,
                         message=f"Mois {i+1}/{nb_mois} — {MOIS_FR[mois]} {annee}…",
                         progress=round(i / nb_mois * 88))
                with _tf.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp_path = tmp.name
                nb_j = generer_template_mois(tp, tmp_path, annee, mois, mode=generation_mode)
                total_jours += nb_j
                wb_tmp = load_workbook(tmp_path)
                ws_src = wb_tmp.active
                ws_dst = wb_out.create_sheet(title=f"{MOIS_FR[mois][:4]} {annee}")
                if generation_mode == 'hide':
                    _copier_feuille_rapide(ws_src, ws_dst)
                else:
                    _copier_feuille(ws_src, ws_dst)
                wb_tmp.close()
                os.unlink(tmp_path)
            _job_set(job_id, message="Sauvegarde du fichier…", progress=94)
            excel_name = f"Templates_{MOIS_FR_UPPER[m_deb]}_{a_deb}_au_{MOIS_FR_UPPER[m_fin]}_{a_fin}.xlsx"
            wb_out.save(os.path.join(wd, excel_name))
            _job_set(job_id, status='done', progress=100, message='Prêt !',
                     fichier=excel_name, session_id=sid,
                     nb_jours=total_jours, nb_mois=nb_mois, format='excel', mois=label)

        elif nb_mois == 1:
            annee, mois = mois_liste[0]
            _job_set(job_id, message=f"Génération {MOIS_FR[mois]} {annee}…", progress=10)
            on = f"{MOIS_FR_UPPER[mois]}_{annee}.xlsx"
            nb_j = generer_template_mois(tp, os.path.join(wd, on), annee, mois, mode=generation_mode)
            _job_set(job_id, status='done', progress=100, message='Prêt !',
                     fichier=on, session_id=sid,
                     nb_jours=nb_j, nb_mois=1, format='xlsx', mois=label)

        else:
            fichiers = []
            for i, (annee, mois) in enumerate(mois_liste):
                _job_set(job_id,
                         message=f"Mois {i+1}/{nb_mois} — {MOIS_FR[mois]} {annee}…",
                         progress=round(i / nb_mois * 88))
                on = f"{MOIS_FR_UPPER[mois]}_{annee}.xlsx"
                nb_j = generer_template_mois(tp, os.path.join(wd, on), annee, mois, mode=generation_mode)
                fichiers.append(on)
                total_jours += nb_j
            _job_set(job_id, message="Compression ZIP…", progress=94)
            zip_name = f"Templates_{MOIS_FR_UPPER[m_deb]}_{a_deb}_au_{MOIS_FR_UPPER[m_fin]}_{a_fin}.zip"
            zip_path = os.path.join(wd, zip_name)
            with _zipfile.ZipFile(zip_path, 'w', _zipfile.ZIP_DEFLATED) as zf:
                for fn in fichiers:
                    zf.write(os.path.join(wd, fn), fn)
            _job_set(job_id, status='done', progress=100, message='Prêt !',
                     fichier=zip_name, session_id=sid,
                     nb_jours=total_jours, nb_mois=nb_mois, format='zip', mois=label)

    except Exception as e:
        import traceback
        _job_set(job_id, status='error', progress=0,
                 message='Erreur', error=str(e), trace=traceback.format_exc())


@app.route('/job/<job_id>')
def job_status(job_id):
    with JOBS_LOCK:
        job = dict(JOBS.get(job_id, {}))
    if not job:
        return jsonify({'error': 'Job introuvable'}), 404
    return jsonify(job)


@app.route('/generer-template-vierge', methods=['POST'])
def generer_template_vierge_route():
    try:
        sid = str(uuid.uuid4())[:8]
        wd  = os.path.join(UPLOAD_FOLDER, sid)
        os.makedirs(wd, exist_ok=True)

        tf = request.files.get('template')
        if not tf:
            return jsonify({'error': 'Template manquant'}), 400
        tp = os.path.join(wd, 'template.xlsx')
        tf.save(tp)

        mois_json = request.form.get('mois_json')
        if mois_json:
            raw = json.loads(mois_json)
            mois_liste = [(int(a), int(m)) for a, m in raw][:12]
        else:
            annee_debut = int(request.form.get('annee_debut', 2026))
            mois_debut  = int(request.form.get('mois_debut',  1))
            annee_fin   = int(request.form.get('annee_fin',   2026))
            mois_fin    = int(request.form.get('mois_fin',    12))
            mois_liste  = []
            a, m = annee_debut, mois_debut
            while (a, m) <= (annee_fin, mois_fin):
                mois_liste.append((a, m))
                m += 1
                if m > 12: m = 1; a += 1
                if len(mois_liste) > 12: break

        if not mois_liste:
            return jsonify({'error': 'Plage de mois invalide'}), 400

        format_sortie   = request.form.get('format', 'zip')
        generation_mode = request.form.get('generation_mode', 'delete')
        if generation_mode not in ('delete', 'hide'):
            generation_mode = 'delete'

        job_id = str(uuid.uuid4())[:12]
        with JOBS_LOCK:
            JOBS[job_id] = {
                'status': 'running', 'progress': 0, 'message': 'Démarrage…',
                'fichier': None, 'session_id': sid, 'error': None, 'trace': None,
                'nb_mois': None, 'nb_jours': None, 'format': None, 'mois': None,
            }

        threading.Thread(
            target=_run_generer_template_vierge,
            args=(job_id, tp, mois_liste, wd, sid, format_sortie, generation_mode),
            daemon=True,
        ).start()

        return jsonify({'job_id': job_id})

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/debug-xml', methods=['GET', 'POST'])
def debug_xml():
    """Route temporaire de debug — inspecte le XML d'un xlsx généré."""
    if request.method == 'GET':
        return '''<form method="post" enctype="multipart/form-data">
            <input type="file" name="template">
            <button type="submit">Analyser</button>
        </form>'''

    import zipfile, io, tempfile
    from xml.etree import ElementTree as ET
    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

    tf = request.files.get('template')
    if not tf:
        return "Template manquant", 400
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tp = tmp.name
    tf.save(tp)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        out = tmp.name
    generer_template_mois(tp, out, 2026, 9, mode='delete')

    result = {}
    with zipfile.ZipFile(out, 'r') as z:
        names = z.namelist()
        result['files'] = names
        # Shared strings
        sst = next((n for n in names if 'sharedStrings' in n), None)
        if sst:
            sst_tree = ET.fromstring(z.read(sst))
            strings = []
            for i, si in enumerate(sst_tree):
                t = si.find(f'{{{NS}}}t')
                strings.append(f"{i}: {t.text if t is not None else '?'}")
            result['sharedStrings'] = strings[:20]
        # Premières lignes de la feuille
        sheet = next((n for n in names if 'worksheets/sheet' in n), None)
        if sheet:
            stree = ET.fromstring(z.read(sheet))
            sd = stree.find(f'{{{NS}}}sheetData')
            rows_info = []
            for row_el in list(sd)[:20]:
                r = row_el.get('r')
                ht = row_el.get('ht')
                cells = []
                for c in row_el:
                    cells.append({'r': c.get('r'), 't': c.get('t'), 'v': (c.find(f'{{{NS}}}v') or ET.Element('x')).text})
                rows_info.append({'r': r, 'ht': ht, 'cells': cells[:3]})
            result['first_rows'] = rows_info

    os.unlink(out)
    return jsonify(result)


@app.route('/fusionner', methods=['POST'])
def fusionner_route():
    sid = str(uuid.uuid4())[:8]
    wd  = os.path.join(UPLOAD_FOLDER, sid)
    os.makedirs(wd, exist_ok=True)
    try:
        fichiers = [f for f in request.files.getlist('fichiers') if f.filename]
        if len(fichiers) < 2:
            return jsonify({'error': 'Au moins 2 fichiers requis'}), 400
        try:
            noms_feuilles = json.loads(request.form.get('noms_json', '[]'))
        except Exception:
            noms_feuilles = []
        sources = [f.read() for f in fichiers]
        result  = fusionner_excels(sources, noms_feuilles or None)
        on = 'Fusion.xlsx'
        op = os.path.join(wd, on)
        with open(op, 'wb') as out:
            out.write(result)
        return jsonify({'fichier': on, 'session_id': sid, 'nb_feuilles': len(sources)})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/telecharger/<session_id>/<filename>')
def telecharger(session_id, filename):
    path = os.path.join(UPLOAD_FOLDER, session_id, filename)
    if not os.path.exists(path): return "Fichier introuvable", 404
    @after_this_request
    def cleanup(r):
        try: shutil.rmtree(os.path.join(UPLOAD_FOLDER, session_id))
        except: pass
        return r
    return send_file(path, as_attachment=True, download_name=filename)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

import sys, io, re, os, openpyxl, calendar, datetime, base64, copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.cell_range import CellRange

with open('app.py', 'r', encoding='utf-8') as f:
    src = f.read()

const_names = ['FERIES', 'JOURS_FR_LIST', 'MOIS_FR_UPPER', 'JOURS_COLS',
               'SLOT_ROWS', 'TEMPLATE_LAST_ROW', 'ALL_MERGE_PAIRS', '_TEMPLATE_VIERGE_B64']

const_block = ""
lines = src.split('\n')
i = 0
while i < len(lines):
    l = lines[i]
    if any(l.startswith(name + ' ') or l.startswith(name + '=') for name in const_names):
        block = [l]
        i += 1
        chars_cont = ('"', "'", ')', ']', '}', '#', '+')
        while i < len(lines) and (
            lines[i].startswith(' ') or lines[i].startswith('\t') or
            (lines[i] and lines[i][0] in chars_cont)
        ):
            block.append(lines[i])
            i += 1
        const_block += '\n'.join(block) + '\n'
    else:
        i += 1

m = re.search(r'^def _appliquer_mois_sur_feuille\(.*?(?=^def |\Z)', src, re.MULTILINE | re.DOTALL)
func_block = m.group(0) if m else ""

exec(compile(const_block + '\n' + func_block, '<app_extract>', 'exec'))

MOIS_TEST = [
    (2026, 9), (2026, 10), (2026, 11), (2026, 12),
    (2027, 1), (2027, 2), (2027, 3), (2027, 4),
    (2027, 5), (2027, 6), (2027, 7), (2027, 8),
]
MOIS_NOM = {9:"SEPTEMBRE",10:"OCTOBRE",11:"NOVEMBRE",12:"DECEMBRE",
            1:"JANVIER",2:"FEVRIER",3:"MARS",4:"AVRIL",
            5:"MAI",6:"JUIN",7:"JUILLET",8:"AOUT"}

print(f"{'Mois':<16} {'nb_jours':>8} {'last_row':>8} {'height':>7} {'border':>8} {'fill':>14} {'OK':>6}")
print("-" * 74)
all_ok = True
for annee, mois in MOIS_TEST:
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(_TEMPLATE_VIERGE_B64)))
    ws = wb.active
    nb = _appliquer_mois_sur_feuille(ws, annee, mois)
    last_row = ws.max_row
    h = ws.row_dimensions[last_row].height if last_row in ws.row_dimensions else None
    cell = ws.cell(last_row, 1)
    bt = cell.border.top.border_style if cell.border else None
    fc = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    ok = bool(h and abs(h - 6.6) < 0.5 and bt == "medium" and fc and "C0C0C0" in fc)
    if not ok:
        all_ok = False
    label = f"{MOIS_NOM[mois]} {annee}"
    ok_str = "OK" if ok else "ERREUR"
    print(f"{label:<16} {nb:>8} {last_row:>8} {str(h or '?'):>7} {str(bt):>8} {str(fc):>14} {ok_str:>6}")

print("-" * 74)
print("RESULTAT:", "TOUT OK" if all_ok else "ERREURS DETECTEES")

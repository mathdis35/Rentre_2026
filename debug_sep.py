import sys, re, io, base64, openpyxl, datetime, calendar, os, copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.cell_range import CellRange

with open("app.py", encoding="utf-8") as f:
    src = f.read()

g = {
    'openpyxl': openpyxl, 'load_workbook': load_workbook, 'Workbook': Workbook,
    'calendar': calendar, 'datetime': datetime,
    'base64': base64, 'io': io, 'copy': copy, 're': re, 'os': os,
    'get_column_letter': get_column_letter,
    'Border': Border, 'Side': Side, 'PatternFill': PatternFill,
    'Font': Font, 'Alignment': Alignment, 'CellRange': CellRange,
}
const_names = [
    'FERIES', 'JOURS_FR_LIST', 'MOIS_FR_UPPER', 'JOURS_COLS',
    'SLOT_ROWS', 'TEMPLATE_LAST_ROW', 'ALL_MERGE_PAIRS', '_TEMPLATE_VIERGE_B64',
]
lines = src.split('\n')
i = 0
const_block_lines = []
chars_cont = set('"\')}]#+')
while i < len(lines):
    l = lines[i]
    if any(l.startswith(name + ' ') or l.startswith(name + '=') for name in const_names):
        const_block_lines.append(l)
        i += 1
        while i < len(lines) and (
            lines[i].startswith(' ') or lines[i].startswith('\t') or
            (lines[i] and lines[i][0] in chars_cont)
        ):
            const_block_lines.append(lines[i])
            i += 1
    else:
        i += 1

exec(compile('\n'.join(const_block_lines), '<consts>', 'exec'), g)

m = re.search(r"^def _appliquer_mois_sur_feuille\(.*?(?=^def |\Z)", src, re.MULTILINE | re.DOTALL)
exec(compile(m.group(0), "<func>", "exec"), g)

_appliquer = g["_appliquer_mois_sur_feuille"]
_b64 = g["_TEMPLATE_VIERGE_B64"]
JOURS_COLS = g["JOURS_COLS"]
SLOT_ROWS = g["SLOT_ROWS"]

ws = openpyxl.load_workbook(io.BytesIO(base64.b64decode(_b64))).active
_appliquer(ws, 2026, 9)  # Septembre, slot_deb=1, nb=22

# Septembre: last_slot_end = SLOT_ROWS[1+22-1]-4+3 = SLOT_ROWS[22]-4+3 = 98-4+3 = 97
# closing_row = 98
print("Septembre - closing_row attendu = 98")
for r in range(94, 102):
    v = ws.cell(r, 2).value
    h = ws.row_dimensions[r].height
    cell = ws._cells.get((r, 2))
    b = cell.border if cell and cell.has_style else None
    bt = b.top.border_style if b else None
    bb = b.bottom.border_style if b else None
    print(f"  row {r}: value={v!r}  h={h}  bt={bt}  bb={bb}")

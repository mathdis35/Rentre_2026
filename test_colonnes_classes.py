"""
Test de détection des colonnes de classes dans le template embarqué.
Colonnes de classes = toutes les colonnes sauf :
  - Colonnes dates (JOURS_COLS)
  - Séparateurs de sections (larges, sans bordure en row 9)
  - Séparateurs entre classes (width <= 1 ou None)
"""
import base64, io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from app import _TEMPLATE_VIERGE_B64, JOURS_COLS

wb = load_workbook(io.BytesIO(base64.b64decode(_TEMPLATE_VIERGE_B64)))
ws = wb.active

# Séparateurs de sections : colonne large (width > 1) sans bordure en row 9
seps_sections = set()
for c in range(1, ws.max_column + 1):
    w = ws.column_dimensions[get_column_letter(c)].width
    if w is None or w <= 1:
        continue
    cell = ws.cell(9, c)
    b = cell.border if cell.has_style else None
    has_border = b and (b.left.border_style or b.right.border_style or b.bottom.border_style)
    if not has_border:
        seps_sections.add(c)

# Séparateurs entre classes : width <= 1 ou None
seps_etroits = set()
for c in range(1, ws.max_column + 1):
    w = ws.column_dimensions[get_column_letter(c)].width
    if w is None or w <= 1:
        seps_etroits.add(c)

# Colonnes de classes = tout le reste
cols_classes = []
for c in range(1, ws.max_column + 1):
    if c in JOURS_COLS:
        continue
    if c in seps_sections:
        continue
    if c in seps_etroits:
        continue
    cols_classes.append(c)

print(f"{'Col':>4} | {'Lettre':>6} | {'Width':>8} | Valeur row 4")
print("-" * 50)
for c in cols_classes:
    lettre = get_column_letter(c)
    w = ws.column_dimensions[lettre].width
    v = ws.cell(4, c).value
    print(f"{c:>4} | {lettre:>6} | {w:>8.2f} | {v}")

print()
print(f"Total colonnes classes détectées : {len(cols_classes)}")
print(f"Lettres : {', '.join(get_column_letter(c) for c in cols_classes)}")
print(f"Largeur max : {max(ws.column_dimensions[get_column_letter(c)].width for c in cols_classes):.4f}")
print(f"Largeur min : {min(ws.column_dimensions[get_column_letter(c)].width for c in cols_classes):.4f}")

# -*- coding: utf-8 -*-
"""
Programme de vérification de cohérence du planning formateurs.

Croise 3 sources :
  1. Tableau matières-formateurs  (heures PRÉVUES, format AFFECTATIONS)
  2. Planning formateurs généré    (heures PLACÉES, lues dans les cases)
  3. Dossier de disponibilités     (1 fichier .xlsx par formateur)

Produit un rapport console :
  A. Heures par formateur : prévu / placé / écart / statut
  B. Respect des disponibilités (un prof n'est jamais placé un créneau où il n'est pas dispo)
  C. Heures manquantes par classe (prévu − placé)
  D. Formateurs du tableau sans fichier de disponibilité

Usage :
    python -X utf8 verifier_coherence.py
    python -X utf8 verifier_coherence.py --planning "Planning_Formateurs_2026_2027.xlsx"
"""
import os, glob, re, datetime, argparse
from collections import defaultdict
from openpyxl import load_workbook

import app  # réutilise les parseurs et la logique de matching de l'application

# ─── Sources par défaut ───────────────────────────────────────────────────────
DEF_PLANNING = 'Planning_Formateurs_2026_2027.xlsx'
DEF_MATIERE_DIR = 'Tableu dispo formateur matière'
DEF_DISPO_DIR = 'Fichier Disponibilités professeurs'

SEUIL_ECART = 20  # tolérance (h) avant de signaler sur/sous-affectation


def trouver_fichier_matiere(dossier):
    fs = glob.glob(os.path.join(dossier, '*.xlsx'))
    if not fs:
        raise FileNotFoundError(f"Aucun tableau matière dans '{dossier}'")
    return fs[0]


def lire_prevu(f_matiere):
    """{formateur: heures} et {classe: heures} prévues depuis le tableau matière."""
    aff = app._auto_parse_formateurs(f_matiere)
    aff.pop('_debug', None)
    par_prof = defaultdict(float)
    par_classe = defaultdict(float)
    for classe, affs in aff.items():
        for a in affs:
            par_prof[a['formateur']] += a['heures']
            par_classe[classe] += a['heures']
    return dict(par_prof), dict(par_classe)


def lire_dispos(dossier):
    """{nom_formateur: {date: {matin1,matin2,pm1,pm2}}}"""
    dispos = {}
    for f in sorted(glob.glob(os.path.join(dossier, '*.xlsx'))):
        d = app.parse_disponibilite(f)
        dispos[d['nom']] = d['dispo']
    return dispos


def _mois_annee_feuille(sn):
    for mk, mnum in app.MOIS_ABBR.items():
        if mk.upper() in sn.upper():
            y = re.search(r'(20\d\d)', sn)
            if y:
                return mnum, int(y.group(1))
    return None, None


def analyser_planning(f_planning):
    """Lit le planning généré (structure 6 lignes/jour) et renvoie :
       - placees_prof   : {prof: heures placées}
       - placees_classe : {classe: heures placées}
       - creneaux       : liste de (date, demi, prof, classe) pour le contrôle dispo
       - sans_prof_classe : {classe: heures 'Pas de prof dispo'}
    """
    wb = load_workbook(f_planning)
    jf = {'lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi'}
    OFF_LABEL = app.MATIN_OFFSETS[-1]
    OFF_NUM = app.PM_OFFSETS[0]
    OFF_PROF_MATIN = app.MATIN_OFFSETS[1]
    OFF_PROF_PM = app.PM_OFFSETS[1]

    placees_prof = defaultdict(float)
    placees_classe = defaultdict(float)
    sans_prof_classe = defaultdict(float)
    creneaux = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        mn, yr = _mois_annee_feuille(sn)
        if not mn:
            continue
        # Colonnes de classes (en-têtes lignes 1-5)
        struct = app.detect_structure(ws)
        classe_de_col = {ci: nom for nom, ci in struct['class_cols'].items()}

        for jcol in app.JOURS_COLS:
            jnext = min([j for j in app.JOURS_COLS if j > jcol], default=ws.max_column + 1)
            for r in range(6, ws.max_row + 1):
                v = ws.cell(r, jcol).value
                if not (isinstance(v, str) and v.strip().lower() in jf):
                    continue
                rl = r - OFF_LABEL
                num = ws.cell(rl + OFF_NUM, jcol).value
                if not isinstance(num, (int, float)):
                    continue
                try:
                    ds = datetime.date(yr, mn, int(num)).strftime('%Y-%m-%d')
                except ValueError:
                    continue
                # Colonnes de classes de cette section
                for c in range(jcol + 1, jnext):
                    classe = classe_de_col.get(c)
                    if not classe:
                        continue
                    for demi, off_prof in [('matin', OFF_PROF_MATIN), ('pm', OFF_PROF_PM)]:
                        pc = ws.cell(rl + off_prof, c).value
                        if not isinstance(pc, str) or not pc.strip():
                            continue
                        if pc.strip() == 'Pas de prof dispo':
                            sans_prof_classe[classe] += 4.0
                            continue
                        profs = [p.strip() for p in pc.split('/') if p.strip()]
                        for prof in profs:
                            h = 4.0 / len(profs)  # demi-journée = 4h, partagée si 2 profs
                            placees_prof[prof] += h
                            placees_classe[classe] += h
                            creneaux.append((ds, demi, prof, classe))
    return dict(placees_prof), dict(placees_classe), creneaux, dict(sans_prof_classe)


def controle_dispos(creneaux, dispos):
    """Vérifie que chaque créneau placé correspond à une disponibilité réelle."""
    ok = non_dispo = sans_fichier = 0
    violations = []
    for ds, demi, prof, classe in creneaux:
        dn = next((n for n in dispos if app.noms_similaires(prof, n)), None)
        if dn is None:
            sans_fichier += 1
            continue
        jd = dispos[dn].get(ds, {})
        dispo = (jd.get('matin1') or jd.get('matin2')) if demi == 'matin' \
            else (jd.get('pm1') or jd.get('pm2'))
        if dispo:
            ok += 1
        else:
            non_dispo += 1
            violations.append((ds, demi, prof, classe))
    return ok, non_dispo, sans_fichier, violations


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--planning', default=DEF_PLANNING)
    ap.add_argument('--matiere-dir', default=DEF_MATIERE_DIR)
    ap.add_argument('--dispo-dir', default=DEF_DISPO_DIR)
    args = ap.parse_args()

    f_matiere = trouver_fichier_matiere(args.matiere_dir)
    print("\n" + "=" * 70)
    print(" VÉRIFICATION DE COHÉRENCE — PLANNING FORMATEURS")
    print("=" * 70)
    print(f"  Planning : {args.planning}")
    print(f"  Tableau  : {os.path.basename(f_matiere)}")
    print(f"  Dispos   : {args.dispo_dir}/")

    prevu_prof, prevu_classe = lire_prevu(f_matiere)
    dispos = lire_dispos(args.dispo_dir)
    placees_prof, placees_classe, creneaux, sans_prof_classe = analyser_planning(args.planning)

    # ── A. Heures par formateur ───────────────────────────────────────────────
    print("\n" + "─" * 70)
    print(" A. HEURES PAR FORMATEUR  (prévu / placé / écart)")
    print("─" * 70)
    print(f"  {'Formateur':<16}{'Prévu':>8}{'Placé':>8}{'Écart':>8}  Statut")
    tous = sorted(set(prevu_prof) | set(placees_prof),
                  key=lambda x: -prevu_prof.get(x, 0))
    tot_prevu = tot_place = 0
    for p in tous:
        pv = prevu_prof.get(p, 0)
        pl = placees_prof.get(p, 0)
        tot_prevu += pv
        tot_place += pl
        ec = pl - pv
        has_dispo = any(app.noms_similaires(p, dn) for dn in dispos)
        if not has_dispo:
            statut = '🔴 AUCUN fichier dispo'
        elif p not in prevu_prof:
            statut = '⚠ absent du tableau'
        elif pl == 0:
            statut = '🔴 jamais placé'
        elif abs(ec) <= SEUIL_ECART:
            statut = '✓'
        elif ec > 0:
            statut = '⚠ sur-affecté'
        else:
            statut = '⚠ sous-affecté'
        print(f"  {p:<16}{pv:>8.0f}{pl:>8.0f}{ec:>+8.0f}  {statut}")
    print(f"  {'TOTAL':<16}{tot_prevu:>8.0f}{tot_place:>8.0f}{tot_place-tot_prevu:>+8.0f}")

    # ── B. Respect des disponibilités ─────────────────────────────────────────
    ok, non_dispo, sans_fichier, violations = controle_dispos(creneaux, dispos)
    total = ok + non_dispo + sans_fichier
    print("\n" + "─" * 70)
    print(" B. RESPECT DES DISPONIBILITÉS")
    print("─" * 70)
    print(f"  Créneaux placés analysés     : {total}")
    pct = (100 * ok / total) if total else 0
    print(f"  ✓ Conformes (prof dispo)     : {ok} ({pct:.1f}%)")
    print(f"  ⚠ Prof PAS dispo ce créneau  : {non_dispo}")
    print(f"  🔴 Sans fichier de dispo      : {sans_fichier}")
    if violations:
        print(f"\n  VIOLATIONS (prof placé alors que non dispo) — {len(violations)} au total :")
        for ds, demi, prof, classe in violations[:30]:
            print(f"    {ds} {demi:<6} {prof:<14} ({classe})")
        if len(violations) > 30:
            print(f"    … et {len(violations) - 30} autres")
    else:
        print("  ✅ Aucune violation : tous les profs placés étaient disponibles.")

    # ── C. Heures manquantes par classe ───────────────────────────────────────
    # Les noms diffèrent entre tableau (ex: "BAC 29") et planning (ex: "BAC PRO 29").
    # On agrège sur les noms du tableau (référence) via noms_similaires.
    print("\n" + "─" * 70)
    print(" C. HEURES MANQUANTES PAR CLASSE  (prévu − placé)")
    print("─" * 70)
    print(f"  {'Classe':<18}{'Prévu':>8}{'Placé':>8}{'Manque':>8}{'Sans prof':>11}")

    ref_classes = list(prevu_classe.keys())  # noms de référence (tableau matière)

    def _aligner(nom):
        m = next((rc for rc in ref_classes if app.noms_similaires(rc, nom)), None)
        return m if m else nom

    placees_aligne = defaultdict(float)
    for cl, h in placees_classe.items():
        placees_aligne[_aligner(cl)] += h
    sansprof_aligne = defaultdict(float)
    for cl, h in sans_prof_classe.items():
        sansprof_aligne[_aligner(cl)] += h

    classes = sorted(set(prevu_classe) | set(placees_aligne),
                     key=lambda x: -(prevu_classe.get(x, 0) - placees_aligne.get(x, 0)))
    for cl in classes:
        pv = prevu_classe.get(cl, 0)
        pl = placees_aligne.get(cl, 0)
        manque = pv - pl
        sp = sansprof_aligne.get(cl, 0)
        flag = '  ⚠' if manque > SEUIL_ECART else ''
        print(f"  {cl:<18}{pv:>8.0f}{pl:>8.0f}{manque:>+8.0f}{sp:>11.0f}{flag}")

    # ── D. Formateurs sans fichier de dispo ───────────────────────────────────
    print("\n" + "─" * 70)
    print(" D. FORMATEURS DU TABLEAU SANS FICHIER DE DISPONIBILITÉ")
    print("─" * 70)
    manquants = [p for p in prevu_prof
                 if not any(app.noms_similaires(p, dn) for dn in dispos)]
    if manquants:
        for p in manquants:
            print(f"  🔴 {p:<16} ({prevu_prof[p]:.0f}h prévues, impossibles à placer)")
    else:
        print("  ✅ Tous les formateurs du tableau ont un fichier de disponibilité.")

    print("\n" + "=" * 70 + "\n")


if __name__ == '__main__':
    main()

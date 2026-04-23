"""
Tests de validation du générateur de template vierge.
Vérifie que chaque mois de l'année scolaire 2026-2027 est généré correctement.

Usage : python test_template_vierge.py

═══════════════════════════════════════════════════════════════════════
PARTIE 1 — Template vierge  (16 tests × 12 mois = 192 tests)
═══════════════════════════════════════════════════════════════════════
  1.  closing_row              - Dernière ligne : h~6.6, border=medium, fill=FFC0C0C0
  2.  pas_de_double_closing    - L'avant-dernière ligne n'a pas aussi border=medium
  3.  premier_jour_row7        - Le premier jour du mois commence bien en row 7
  4.  premier_jour_correct     - Label et numéro en row 7 = vrai 1er jour lun-ven du mois
  5.  dernier_jour_correct     - Le dernier slot utilisé = vrai dernier jour lun-ven du mois
  6.  nb_jours                 - Nombre de jours retourné = nombre de jours lun-ven réels
  7.  pas_de_lignes_fantomes   - Aucune ligne au-delà de la closing row
  8.  pas_de_slot_vide_debut   - Aucun slot vide entre row 6 et le premier jour (row 7)
  9.  pas_de_slot_vide_fin     - Aucun label de jour après le dernier jour utilisé
  10. continuite_des_jours     - Les jours s'enchaînent sans trou ni saut
  11. titre_mois_correct       - Row 1 contient le bon mois (pas "SEPTEMBRE" résiduel)
  12. dates_toutes_sections    - Dates présentes dans les 4 sections pour chaque jour
  13. pas_de_date_residuelle   - Aucune date d'un autre mois ne subsiste
  14. hauteurs_slots           - Les lignes du tableau ont des hauteurs cohérentes (~5-30pt)
  15. merges_preserves         - Les fusions ALL_MERGE_PAIRS sont présentes sur les rows actives
  16. separateurs_semaine      - Les séparateurs gris entre semaines sont bien présents

═══════════════════════════════════════════════════════════════════════
PARTIE 2 — Parsers, moteur d'assignation, routes Flask
═══════════════════════════════════════════════════════════════════════
  find_month_num()             - Parsing des noms de mois (abbréviations, sans année…)
  noms_similaires()            - Matching de noms avec/sans numéros, casse insensible
  is_available()               - Valeurs reconnues comme disponible (X, OUI, ✓, 4…)
  parse_planning_xlsx()        - Fichier xlsx synthétique : nom, jours bleus, week-end exclu
  parse_disponibilite()        - Fichier dispo synthétique : nom, matin/pm, week-end ignoré
  parse_tableau_formateurs_v2()- ACTIF=NON ignoré, heures=0 ignoré, tri priorité, erreurs
  _auto_parse_formateurs()     - Détection v2 si feuille AFFECTATIONS, fallback v1 sinon
  assigner()                   - Assignation, indisponible → ⚠️, sans affectation → ?,
                                 alternance matin/pm, heures épuisées, noms similaires
  Routes Flask                 - /ping, /, /generer-template-vierge, /generer-template,
                                 /generer  (via client de test Flask)
"""

import sys, io, re, os, openpyxl, calendar, datetime, base64, copy, tempfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.cell_range import CellRange

# ---------------------------------------------------------------------------
# Chargement de app.py sans instancier Flask (pour la partie 1)
# ---------------------------------------------------------------------------

with open('app.py', 'r', encoding='utf-8') as f:
    src = f.read()

from openpyxl import load_workbook, Workbook
g = {
    'openpyxl': openpyxl, 'load_workbook': load_workbook, 'Workbook': Workbook,
    'calendar': calendar, 'datetime': datetime,
    'base64': base64, 'io': io, 'copy': copy, 're': re, 'os': os,
    'get_column_letter': get_column_letter,
    'Border': Border, 'Side': Side, 'PatternFill': PatternFill,
    'Font': Font, 'Alignment': Alignment, 'CellRange': CellRange,
}

const_names = [
    'FERIES', 'JOURS_FR_LIST', 'MOIS_FR_UPPER', 'JOURS_COLS',
    'SLOT_ROWS', 'TEMPLATE_LAST_ROW', 'ALL_MERGE_PAIRS', '_TEMPLATE_VIERGE_B64',
]
lines = src.split('\n')
i = 0
const_block_lines = []
chars_cont = set('"\')}]#+')
while i < len(lines):
    l = lines[i]
    if any(l.startswith(name + ' ') or l.startswith(name + '=') for name in const_names):
        const_block_lines.append(l)
        i += 1
        while i < len(lines) and (
            lines[i].startswith(' ') or lines[i].startswith('\t') or
            (lines[i] and lines[i][0] in chars_cont)
        ):
            const_block_lines.append(lines[i])
            i += 1
    else:
        i += 1

exec(compile('\n'.join(const_block_lines), '<consts>', 'exec'), g)

m = re.search(r'^def _appliquer_mois_sur_feuille\(.*?(?=^def |\Z)', src, re.MULTILINE | re.DOTALL)
if not m:
    print("ERREUR : fonction _appliquer_mois_sur_feuille introuvable dans app.py")
    sys.exit(1)
exec(compile(m.group(0), '<func>', 'exec'), g)

_appliquer      = g['_appliquer_mois_sur_feuille']
_b64            = g['_TEMPLATE_VIERGE_B64']
SLOT_ROWS       = g['SLOT_ROWS']
JOURS_COLS      = g['JOURS_COLS']
ALL_MERGE_PAIRS = g['ALL_MERGE_PAIRS']
MOIS_FR_UPPER   = g['MOIS_FR_UPPER']
JOURS_FR_LIST   = g['JOURS_FR_LIST']

# ---------------------------------------------------------------------------
# Mois à tester (année scolaire 2026-2027)
# ---------------------------------------------------------------------------

MOIS_TEST = [
    (2026,  9), (2026, 10), (2026, 11), (2026, 12),
    (2027,  1), (2027,  2), (2027,  3), (2027,  4),
    (2027,  5), (2027,  6), (2027,  7), (2027,  8),
]

MOIS_NOM = {
    9: "Septembre", 10: "Octobre",  11: "Novembre", 12: "Decembre",
    1: "Janvier",   2: "Fevrier",   3: "Mars",      4: "Avril",
    5: "Mai",       6: "Juin",      7: "Juillet",   8: "Aout",
}

JOURS_FR = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_ws():
    return openpyxl.load_workbook(io.BytesIO(base64.b64decode(_b64))).active

def jours_lun_ven(annee, mois):
    return [
        datetime.date(annee, mois, j)
        for j in range(1, calendar.monthrange(annee, mois)[1] + 1)
        if datetime.date(annee, mois, j).weekday() < 5
    ]

def labels_dans_ws(ws):
    result = []
    for r in range(6, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v in JOURS_FR:
            result.append((r, v))
    return result

def nums_dans_ws(ws):
    result = []
    for r in range(6, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, int):
            result.append((r, v))
    return result

# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_closing_row(annee, mois, ws, nb):
    """Dernière ligne : h~6.6, border_top=medium, fill=FFC0C0C0."""
    last_row = ws.max_row
    rd = ws.row_dimensions
    h  = rd[last_row].height if last_row in rd else None
    c2 = ws.cell(last_row, 2)
    bt = c2.border.top.border_style if c2.border else None
    fc = c2.fill.fgColor.rgb if c2.fill and c2.fill.fgColor else None
    assert h is not None and abs(h - 6.6) < 0.5, f"h={h} (attendu ~6.6)"
    assert bt == 'medium', f"border_top={bt!r} (attendu 'medium')"
    assert fc == 'FFC0C0C0', f"fill={fc!r} (attendu 'FFC0C0C0')"


def test_pas_de_double_closing(annee, mois, ws, nb):
    """Seule max_row doit avoir fill=FFC0C0C0 — ni max_row-1 ni max_row+1."""
    last_row = ws.max_row
    for r in [last_row - 1, last_row + 1]:
        cell = ws.cell(r, 2)
        fc = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
        assert fc != 'FFC0C0C0', \
            f"double closing : row {r} a aussi fill=FFC0C0C0 (closing={last_row})"


def test_premier_jour_row7(annee, mois, ws, nb):
    """Le premier label de jour doit apparaître en row 7 (col 2)."""
    first_row = None
    for r in range(6, 15):
        v = ws.cell(r, 2).value
        if v is not None:
            first_row = r
            break
    assert first_row == 7, f"premier jour en row {first_row} (attendu row 7)"


def test_premier_jour_correct(annee, mois, ws, nb):
    """Row 7 = label du 1er jour lun-ven, row 8 = son numéro."""
    jours = jours_lun_ven(annee, mois)
    premier = jours[0]
    label_attendu = JOURS_FR[premier.weekday()]
    num_attendu   = premier.day
    label_ws = ws.cell(7, 2).value
    num_ws   = ws.cell(8, 2).value
    assert label_ws == label_attendu, \
        f"label={label_ws!r} (attendu {label_attendu!r})"
    assert num_ws == num_attendu, \
        f"num={num_ws!r} (attendu {num_attendu!r})"


def test_dernier_jour_correct(annee, mois, ws, nb):
    """Le dernier label/num écrits correspondent au dernier jour lun-ven du mois."""
    jours   = jours_lun_ven(annee, mois)
    dernier = jours[-1]
    label_attendu = JOURS_FR[dernier.weekday()]
    num_attendu   = dernier.day
    labels = labels_dans_ws(ws)
    nums   = nums_dans_ws(ws)
    assert labels, "aucun label de jour trouve dans le ws"
    assert nums,   "aucun numero de jour trouve dans le ws"
    assert labels[-1][1] == label_attendu, \
        f"dernier label={labels[-1][1]!r} (attendu {label_attendu!r})"
    assert nums[-1][1] == num_attendu, \
        f"dernier num={nums[-1][1]!r} (attendu {num_attendu!r})"


def test_nb_jours(annee, mois, ws, nb):
    """Nombre de jours retourné = jours lun-ven réels du mois."""
    attendu = len(jours_lun_ven(annee, mois))
    assert nb == attendu, f"nb={nb} (attendu {attendu})"


def test_pas_de_lignes_fantomes(annee, mois, ws, nb):
    """Aucune ligne (hauteur ou valeur) au-delà de la closing row."""
    last_row = ws.max_row
    rd = ws.row_dimensions
    for r in range(last_row + 1, last_row + 10):
        h_after = rd[r].height if r in rd else None
        v_after = ws.cell(r, 2).value
        assert h_after is None and v_after is None, \
            f"ligne fantome en row {r} (h={h_after}, val={v_after!r})"


def test_pas_de_slot_vide_debut(annee, mois, ws, nb):
    """Row 6 col 2 doit être vide (pas de slot parasite avant row 7)."""
    v6 = ws.cell(6, 2).value
    assert v6 is None, f"row 6 col 2 contient {v6!r} (devrait etre vide)"


def test_pas_de_slot_vide_fin(annee, mois, ws, nb):
    """Nombre de labels dans le ws = nombre de jours du mois (pas de labels orphelins)."""
    labels = labels_dans_ws(ws)
    assert len(labels) == nb, \
        f"{len(labels)} labels trouves (attendu {nb})"


def test_continuite_des_jours(annee, mois, ws, nb):
    """Les jours s'enchaînent sans trou : label et numéro corrects à chaque slot."""
    jours  = jours_lun_ven(annee, mois)
    labels = labels_dans_ws(ws)
    nums   = nums_dans_ws(ws)
    assert len(labels) == len(jours), \
        f"{len(labels)} labels vs {len(jours)} jours attendus"
    for idx, (d, (row_l, label_ws), (row_n, num_ws)) in enumerate(
            zip(jours, labels, nums)):
        label_att = JOURS_FR[d.weekday()]
        assert label_ws == label_att, \
            f"slot {idx}: label={label_ws!r} (attendu {label_att!r}) en row {row_l}"
        assert num_ws == d.day, \
            f"slot {idx}: num={num_ws!r} (attendu {d.day}) en row {row_n}"


def test_titre_mois_correct(annee, mois, ws, nb):
    """Row 1 doit contenir le nom du mois correct, pas 'SEPTEMBRE' résiduel."""
    nom_attendu = MOIS_FR_UPPER[mois]
    found = False
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and nom_attendu in v.upper():
            found = True
            break
    assert found, f"titre '{nom_attendu}' introuvable en row 1"
    if mois != 9:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if isinstance(v, str) and 'SEPTEMBRE' in v.upper():
                assert False, \
                    f"residuel 'SEPTEMBRE' trouve en row 1 col {c} : {v!r}"


def test_dates_toutes_sections(annee, mois, ws, nb):
    """Chaque jour est présent (label + numéro) dans les 4 sections (JOURS_COLS)."""
    labels = labels_dans_ws(ws)
    for row_l, _ in labels:
        for col in JOURS_COLS:
            v = ws.cell(row_l, col).value
            assert v is not None, \
                f"label manquant en row {row_l} col {col}"
        for col in JOURS_COLS:
            v = ws.cell(row_l + 1, col).value
            assert v is not None, \
                f"numero manquant en row {row_l+1} col {col}"


def test_pas_de_date_residuelle(annee, mois, ws, nb):
    """Aucun numéro de jour ne correspond à un jour hors du mois courant."""
    jours = jours_lun_ven(annee, mois)
    nums_attendus = {d.day for d in jours}
    nums = nums_dans_ws(ws)
    for row_n, num in nums:
        assert num in nums_attendus, \
            f"numero {num} en row {row_n} n'appartient pas au mois {MOIS_NOM[mois]} {annee}"


def test_hauteurs_slots(annee, mois, ws, nb):
    """Les lignes des slots actifs ont une hauteur entre 5 et 30 pt."""
    labels = labels_dans_ws(ws)
    rd = ws.row_dimensions
    for row_l, label in labels:
        for r in range(row_l, row_l + 6):
            h = rd[r].height if r in rd else None
            if h is not None:
                assert 5 <= h <= 30, \
                    f"hauteur anormale row {r} : h={h} (slot '{label}')"


def test_merges_preserves(annee, mois, ws, nb):
    """Les fusions ALL_MERGE_PAIRS sont présentes sur les rows des slots actifs."""
    labels = labels_dans_ws(ws)
    merged = {(mg.min_row, mg.min_col, mg.max_col)
              for mg in ws.merged_cells.ranges
              if mg.min_row == mg.max_row}
    for row_l, label in labels:
        for c1, c2 in ALL_MERGE_PAIRS[:3]:
            key = (row_l, c1, c2)
            assert key in merged, \
                f"fusion manquante {get_column_letter(c1)}{row_l}:{get_column_letter(c2)}{row_l} (slot '{label}')"


def test_separateurs_semaine(annee, mois, ws, nb):
    """
    Les séparateurs inter-semaines (ligne grise entre vendredi et lundi suivant)
    doivent être présents entre chaque groupe de 5 jours consécutifs.
    Un séparateur = 1 ligne avec border_top=medium après le slot vendredi.
    """
    labels = labels_dans_ws(ws)
    # Trouver les positions des vendredis (dernier jour de chaque semaine)
    for idx, (row_l, label) in enumerate(labels):
        if label == 'Vendredi' and idx + 1 < len(labels):
            # La row après le slot vendredi (row_l + 6) doit être un séparateur
            sep_row = row_l + 6
            next_slot_row = labels[idx + 1][0]
            # Il doit y avoir exactement 1 ligne entre fin du slot vendredi et début du lundi
            assert next_slot_row == sep_row + 1, \
                f"separateur manquant apres Vendredi row {row_l} : " \
                f"slot suivant en row {next_slot_row} (attendu {sep_row + 1})"

def test_pas_de_lignes_apres_closing(annee, mois, ws, nb):
    """
    Aucune cellule avec style visible (fill coloré ou border) ne doit exister
    au-delà de la closing row. Détecte les lignes fantômes résiduelles.
    """
    last_row = ws.max_row
    TRANSPARENT = {'00000000', '000000', 'FF000000', None}
    for (r, c), cell in ws._cells.items():
        if r <= last_row:
            continue
        if not cell.has_style:
            continue
        fi = cell.fill
        b  = cell.border
        has_fill   = fi and fi.fill_type not in (None, 'none') and str(fi.fgColor.rgb) not in TRANSPARENT
        has_border = b and any(getattr(b, s).border_style for s in ('left','right','top','bottom'))
        assert not has_fill and not has_border, \
            f"cellule fantome avec style en row {r} col {c} (closing={last_row}) : fill={fi.fgColor.rgb if fi else None} border={b}"


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

TOUS_LES_TESTS = [
    test_closing_row,
    test_pas_de_double_closing,
    test_premier_jour_row7,
    test_premier_jour_correct,
    test_dernier_jour_correct,
    test_nb_jours,
    test_pas_de_lignes_fantomes,
    test_pas_de_slot_vide_debut,
    test_pas_de_slot_vide_fin,
    test_continuite_des_jours,
    test_titre_mois_correct,
    test_dates_toutes_sections,
    test_pas_de_date_residuelle,
    test_hauteurs_slots,
    test_merges_preserves,
    test_separateurs_semaine,
    test_pas_de_lignes_apres_closing,
]


def run_all():
    total  = 0
    passed = 0
    failed = 0
    errors = []

    print(f"\n{'Mois':<18} {'nb_j':>5}  {'Resultat'}")
    print("-" * 55)

    for annee, mois in MOIS_TEST:
        ws      = load_ws()
        nb      = _appliquer(ws, annee, mois)
        label   = f"{MOIS_NOM[mois]} {annee}"
        mois_ok = True

        for fn in TOUS_LES_TESTS:
            total += 1
            try:
                fn(annee, mois, ws, nb)
                passed += 1
            except AssertionError as e:
                failed += 1
                mois_ok = False
                errors.append(f"  [{label}] {fn.__name__}: {e}")

        print(f"{label:<18} {nb:>5}  {'OK' if mois_ok else 'ECHEC'}")

    print("-" * 55)
    print(f"\n{len(MOIS_TEST)} mois x {len(TOUS_LES_TESTS)} tests = {total} au total")
    print(f"RESULTAT : {passed}/{total} passes", end="")
    if failed:
        print(f", {failed} echec(s)\n")
        for e in errors:
            print(e)
    else:
        print(" - TOUT OK")

    return failed == 0


# ===========================================================================
# PARTIE 2 — Parsers, moteur d'assignation, routes Flask
# ===========================================================================

# Import direct de app.py (nécessite Flask et xlrd installés)
try:
    import app as _app
    _APP_OK = True
except Exception as _e:
    _APP_OK = False
    print(f"\n[WARN] Impossible d'importer app.py : {_e}")
    print("       Les tests de la partie 2 seront ignorés.\n")


# ── Helpers fichiers synthétiques ────────────────────────────────────────────

def _tmp_xlsx(wb):
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def _make_planning_wb(jours_bleus):
    """jours_bleus : liste de (annee, mois, jour) marqués en bleu dans col cj=4, jour en col cd=5."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "BAC PRO COMMERCE 28"
    ws.cell(5, 3).value = "Septembre 2026"  # mois en col 3 → cj=4, cd=5
    blue = PatternFill(fill_type="solid", fgColor="FFA6CAF0")
    for r, (_, _, jour) in enumerate(jours_bleus, start=6):
        ws.cell(r, 4).fill = blue   # cj = col_mois + 1 = 4
        ws.cell(r, 5).value = jour  # cd = col_mois + 2 = 5
    return wb


def _make_dispo_wb(nom_formateur, disponibilites):
    """disponibilites : liste de (annee, mois, jour, matin:bool, pm:bool)."""
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = "Nom :"
    ws.cell(1, 2).value = nom_formateur
    ws.cell(3, 3).value = datetime.datetime(2026, 9, 1, 0, 0)
    jours_abbr = ['lun', 'mar', 'mer', 'jeu', 'ven', 'sam', 'dim']
    row = 5
    for annee, mois, jour, matin, pm in disponibilites:
        d = datetime.date(annee, mois, jour)
        ws.cell(row, 1).value = jours_abbr[d.weekday()]
        ws.cell(row, 2).value = jour
        ws.cell(row, 3).value = "X" if matin else None
        ws.cell(row, 4).value = "X" if pm else None
        row += 1
    return wb


def _make_aff_wb(rows, sheet_name="AFFECTATIONS"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for c, h in enumerate(["CLASSE", "FORMATEUR", "MATIERE", "HEURES_ANNEE", "PRIORITE", "ACTIF"], 1):
        ws.cell(1, c).value = h
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(r, c).value = val
    return wb


# ── Définition des tests partie 2 ────────────────────────────────────────────

P2_TESTS = []  # liste de (nom, callable) → callable() lève AssertionError ou retourne


def _p2(name):
    def decorator(fn):
        P2_TESTS.append((name, fn))
        return fn
    return decorator


# find_month_num ----------------------------------------------------------

@_p2("find_month_num: septembre 2026")
def _():
    mn, yr = _app.find_month_num("Septembre 2026")
    assert mn == 9 and yr == 2026, f"mn={mn} yr={yr}"

@_p2("find_month_num: JANVIER 2027")
def _():
    mn, yr = _app.find_month_num("JANVIER 2027")
    assert mn == 1 and yr == 2027

@_p2("find_month_num: abréviation sept.")
def _():
    mn, _ = _app.find_month_num("sept. 2026")
    assert mn == 9

@_p2("find_month_num: sans année → yr None")
def _():
    mn, yr = _app.find_month_num("mars")
    assert mn is not None and yr is None

@_p2("find_month_num: texte inconnu → None,None")
def _():
    mn, yr = _app.find_month_num("bonjour")
    assert mn is None and yr is None

@_p2("find_month_num: février")
def _():
    mn, _ = _app.find_month_num("Février 2027")
    assert mn == 2

@_p2("find_month_num: août")
def _():
    mn, _ = _app.find_month_num("août 2027")
    assert mn == 8


# noms_similaires ----------------------------------------------------------

@_p2("noms_similaires: classes identiques")
def _():
    assert _app.noms_similaires("BTS MCO 73", "BTS MCO 73")

@_p2("noms_similaires: classes même num même mot")
def _():
    assert _app.noms_similaires("BAC PRO 28", "BAC PRO COMMERCE 28")

@_p2("noms_similaires: classes numéros différents → False")
def _():
    assert not _app.noms_similaires("BTS MCO 73", "BTS MCO 74")

@_p2("noms_similaires: même num mot différent → False")
def _():
    assert not _app.noms_similaires("BTS MCO 73", "BAC PRO 73")

@_p2("noms_similaires: formateurs identiques")
def _():
    assert _app.noms_similaires("Jean Dupont", "Jean Dupont")

@_p2("noms_similaires: formateurs 2 mots communs")
def _():
    assert _app.noms_similaires("Marie Claire Dupont", "Claire Dupont")

@_p2("noms_similaires: formateurs 1 mot court")
def _():
    assert _app.noms_similaires("Dupont", "Dupont Martin")

@_p2("noms_similaires: formateurs aucun commun → False")
def _():
    assert not _app.noms_similaires("Jean Dupont", "Pierre Martin")

@_p2("noms_similaires: insensible à la casse")
def _():
    assert _app.noms_similaires("jean dupont", "JEAN DUPONT")


# is_available -------------------------------------------------------------

@_p2("is_available: X")
def _():
    assert _app.is_available("X")

@_p2("is_available: x minuscule")
def _():
    assert _app.is_available("x")

@_p2("is_available: OUI")
def _():
    assert _app.is_available("OUI")

@_p2("is_available: O")
def _():
    assert _app.is_available("O")

@_p2("is_available: ✓")
def _():
    assert _app.is_available("✓")

@_p2("is_available: 4")
def _():
    assert _app.is_available("4")

@_p2("is_available: None → False")
def _():
    assert not _app.is_available(None)

@_p2("is_available: vide → False")
def _():
    assert not _app.is_available("")

@_p2("is_available: NON → False")
def _():
    assert not _app.is_available("NON")

@_p2("is_available: tiret → False")
def _():
    assert not _app.is_available("-")


# parse_planning_xlsx ------------------------------------------------------

@_p2("parse_planning_xlsx: nom extrait de row 1")
def _():
    path = _tmp_xlsx(_make_planning_wb([(2026, 9, 7)]))
    try:
        res = _app.parse_planning_xlsx(path)
        assert res['nom'] == "BAC PRO COMMERCE 28", f"nom={res['nom']!r}"
    finally:
        os.unlink(path)

@_p2("parse_planning_xlsx: lundi 7 sept détecté")
def _():
    path = _tmp_xlsx(_make_planning_wb([(2026, 9, 7)]))
    try:
        res = _app.parse_planning_xlsx(path)
        assert "2026-09-07" in res['jours']
    finally:
        os.unlink(path)

@_p2("parse_planning_xlsx: samedi exclu")
def _():
    path = _tmp_xlsx(_make_planning_wb([(2026, 9, 5)]))
    try:
        res = _app.parse_planning_xlsx(path)
        assert "2026-09-05" not in res['jours']
    finally:
        os.unlink(path)

@_p2("parse_planning_xlsx: plusieurs jours")
def _():
    path = _tmp_xlsx(_make_planning_wb([(2026, 9, 7), (2026, 9, 8), (2026, 9, 9)]))
    try:
        res = _app.parse_planning_xlsx(path)
        assert len(res['jours']) == 3
    finally:
        os.unlink(path)

@_p2("parse_planning_xlsx: fichier vide → jours=[]")
def _():
    path = _tmp_xlsx(Workbook())
    try:
        res = _app.parse_planning_xlsx(path)
        assert res['jours'] == []
    finally:
        os.unlink(path)

@_p2("parse_planning_xlsx: jours triés")
def _():
    path = _tmp_xlsx(_make_planning_wb([(2026, 9, 9), (2026, 9, 7), (2026, 9, 8)]))
    try:
        res = _app.parse_planning_xlsx(path)
        assert res['jours'] == sorted(res['jours'])
    finally:
        os.unlink(path)


# parse_disponibilite ------------------------------------------------------

@_p2("parse_disponibilite: nom extrait")
def _():
    path = _tmp_xlsx(_make_dispo_wb("Dupont Jean", [(2026, 9, 7, True, False)]))
    try:
        res = _app.parse_disponibilite(path)
        assert res['nom'] == "Dupont Jean", f"nom={res['nom']!r}"
    finally:
        os.unlink(path)

@_p2("parse_disponibilite: dispo matin")
def _():
    path = _tmp_xlsx(_make_dispo_wb("Test", [(2026, 9, 7, True, False)]))
    try:
        res = _app.parse_disponibilite(path)
        assert "2026-09-07" in res['dispo']
        assert res['dispo']["2026-09-07"]['matin'] is True
        assert res['dispo']["2026-09-07"]['pm'] is False
    finally:
        os.unlink(path)

@_p2("parse_disponibilite: dispo pm")
def _():
    path = _tmp_xlsx(_make_dispo_wb("Test", [(2026, 9, 8, False, True)]))
    try:
        res = _app.parse_disponibilite(path)
        assert "2026-09-08" in res['dispo']
        assert res['dispo']["2026-09-08"]['pm'] is True
    finally:
        os.unlink(path)

@_p2("parse_disponibilite: samedi ignoré")
def _():
    path = _tmp_xlsx(_make_dispo_wb("Test", [(2026, 9, 5, True, True)]))
    try:
        res = _app.parse_disponibilite(path)
        assert "2026-09-05" not in res['dispo']
    finally:
        os.unlink(path)

@_p2("parse_disponibilite: sans X → absent")
def _():
    path = _tmp_xlsx(_make_dispo_wb("Test", [(2026, 9, 7, False, False)]))
    try:
        res = _app.parse_disponibilite(path)
        assert "2026-09-07" not in res['dispo']
    finally:
        os.unlink(path)

@_p2("parse_disponibilite: fichier vide → dispo={}")
def _():
    path = _tmp_xlsx(Workbook())
    try:
        res = _app.parse_disponibilite(path)
        assert res['dispo'] == {}
    finally:
        os.unlink(path)


# parse_tableau_formateurs_v2 ---------------------------------------------

@_p2("parse_v2: parse basique")
def _():
    path = _tmp_xlsx(_make_aff_wb([("BTS MCO 73", "Jean Dupont", "Gestion", 100, 1, "OUI")]))
    try:
        res = _app.parse_tableau_formateurs_v2(path)
        assert "BTS MCO 73" in res
        assert res["BTS MCO 73"][0]['formateur'] == "Jean Dupont"
        assert res["BTS MCO 73"][0]['heures'] == 100.0
    finally:
        os.unlink(path)

@_p2("parse_v2: heures_faites initialisées à 0")
def _():
    path = _tmp_xlsx(_make_aff_wb([("BTS MCO 73", "Jean Dupont", "Gestion", 100, 1, "OUI")]))
    try:
        res = _app.parse_tableau_formateurs_v2(path)
        assert res["BTS MCO 73"][0]['heures_faites'] == 0
    finally:
        os.unlink(path)

@_p2("parse_v2: ACTIF=NON ignoré")
def _():
    path = _tmp_xlsx(_make_aff_wb([
        ("BTS MCO 73", "Jean Dupont", "Gestion", 100, 1, "NON"),
        ("BTS MCO 73", "Marie Martin", "Marketing", 80, 1, "OUI"),
    ]))
    try:
        res = _app.parse_tableau_formateurs_v2(path)
        noms = [a['formateur'] for a in res.get("BTS MCO 73", [])]
        assert "Jean Dupont" not in noms
        assert "Marie Martin" in noms
    finally:
        os.unlink(path)

@_p2("parse_v2: heures=0 ignoré")
def _():
    path = _tmp_xlsx(_make_aff_wb([("BTS MCO 73", "Jean Dupont", "Gestion", 0, 1, "OUI")]))
    try:
        res = _app.parse_tableau_formateurs_v2(path)
        assert len(res.get("BTS MCO 73", [])) == 0
    finally:
        os.unlink(path)

@_p2("parse_v2: tri par priorité")
def _():
    path = _tmp_xlsx(_make_aff_wb([
        ("BTS MCO 73", "Remplacant", "Gestion", 50, 2, "OUI"),
        ("BTS MCO 73", "Principal", "Gestion", 100, 1, "OUI"),
    ]))
    try:
        res = _app.parse_tableau_formateurs_v2(path)
        noms = [a['formateur'] for a in res["BTS MCO 73"]]
        assert noms[0] == "Principal" and noms[1] == "Remplacant"
    finally:
        os.unlink(path)

@_p2("parse_v2: plusieurs classes")
def _():
    path = _tmp_xlsx(_make_aff_wb([
        ("BTS MCO 73", "Jean Dupont", "Gestion", 100, 1, "OUI"),
        ("BAC PRO 28", "Marie Martin", "Vente", 80, 1, "OUI"),
    ]))
    try:
        res = _app.parse_tableau_formateurs_v2(path)
        assert "BTS MCO 73" in res and "BAC PRO 28" in res
    finally:
        os.unlink(path)

@_p2("parse_v2: entête manquant → ValueError")
def _():
    wb = Workbook(); ws = wb.active; ws.title = "AFFECTATIONS"
    ws.cell(1, 1).value = "pas_un_entete"
    path = _tmp_xlsx(wb)
    try:
        raised = False
        try:
            _app.parse_tableau_formateurs_v2(path)
        except ValueError:
            raised = True
        assert raised, "ValueError attendue"
    finally:
        os.unlink(path)

@_p2("parse_v2: _debug présent dans le résultat")
def _():
    path = _tmp_xlsx(_make_aff_wb([("BTS MCO 73", "Jean Dupont", "Gestion", 100, 1, "OUI")]))
    try:
        res = _app.parse_tableau_formateurs_v2(path)
        assert '_debug' in res and res['_debug']['classes'] >= 1
    finally:
        os.unlink(path)


# _auto_parse_formateurs ---------------------------------------------------

@_p2("auto_parse: détecte v2 si feuille AFFECTATIONS")
def _():
    wb = Workbook(); ws = wb.active; ws.title = "AFFECTATIONS"
    for c, h in enumerate(["CLASSE", "FORMATEUR", "MATIERE", "HEURES_ANNEE"], 1):
        ws.cell(1, c).value = h
    ws.cell(2, 1).value = "BTS MCO 73"; ws.cell(2, 2).value = "Dupont"
    ws.cell(2, 3).value = "Gestion";    ws.cell(2, 4).value = 100
    path = _tmp_xlsx(wb)
    try:
        res = _app._auto_parse_formateurs(path)
        assert "BTS MCO 73" in res
    finally:
        os.unlink(path)

@_p2("auto_parse: fallback v1 si pas de feuille AFFECTATIONS")
def _():
    wb = Workbook(); wb.active.title = "Sheet1"
    path = _tmp_xlsx(wb)
    try:
        res = _app._auto_parse_formateurs(path)
        assert isinstance(res, dict)
    finally:
        os.unlink(path)


# assigner -----------------------------------------------------------------

def _planning(nom, jours):
    return {'nom': nom, 'jours': jours}

def _dispo(nom, dispo_dict):
    return {'nom': nom, 'dispo': dispo_dict}

def _aff_simple(classe, formateur, matiere="Gestion", heures=100):
    return {classe: [{'formateur': formateur, 'matiere': matiere,
                      'heures': heures, 'heures_faites': 0, 'priorite': 1}]}


@_p2("assigner: assignation simple")
def _():
    res, stats, _ = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07"])],
        [_dispo("Jean Dupont", {"2026-09-07": {"matin": True, "pm": False}})],
        _aff_simple("BTS MCO 73", "Jean Dupont"),
    )
    assert res["2026-09-07"]["BTS MCO 73"]["formateur"] == "Jean Dupont"
    assert stats['assigned'] == 1 and stats['warn'] == 0

@_p2("assigner: formateur indisponible → ⚠️")
def _():
    res, stats, _ = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07"])],
        [_dispo("Jean Dupont", {"2026-09-07": {"matin": False, "pm": False}})],
        _aff_simple("BTS MCO 73", "Jean Dupont"),
    )
    assert res["2026-09-07"]["BTS MCO 73"]["formateur"] == "⚠️"
    assert stats['warn'] == 1

@_p2("assigner: classe sans affectation → ?")
def _():
    res, _, _ = _app.assigner(
        [_planning("BTS INCONNU", ["2026-09-07"])], [], {}
    )
    assert res["2026-09-07"]["BTS INCONNU"]["formateur"] == "?"

@_p2("assigner: alternance matin/pm")
def _():
    dispo_jours = {
        "2026-09-07": {"matin": True, "pm": True},
        "2026-09-08": {"matin": True, "pm": True},
    }
    res, _, _ = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07", "2026-09-08"])],
        [_dispo("Jean Dupont", dispo_jours)],
        _aff_simple("BTS MCO 73", "Jean Dupont"),
    )
    s1 = res["2026-09-07"]["BTS MCO 73"]["slot"]
    s2 = res["2026-09-08"]["BTS MCO 73"]["slot"]
    assert s1 != s2, f"pas d'alternance : s1={s1} s2={s2}"

@_p2("assigner: heures comptabilisées (+4 par jour)")
def _():
    _, _, heures = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07"])],
        [_dispo("Jean Dupont", {"2026-09-07": {"matin": True, "pm": False}})],
        _aff_simple("BTS MCO 73", "Jean Dupont"),
    )
    assert heures["Jean Dupont"]["Gestion"] == 4

@_p2("assigner: formateur sans fichier dispo → ⚠️")
def _():
    res, stats, _ = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07"])],
        [],  # aucun fichier dispo
        _aff_simple("BTS MCO 73", "Jean Dupont"),
    )
    assert res["2026-09-07"]["BTS MCO 73"]["formateur"] == "⚠️"

@_p2("assigner: matching par noms similaires")
def _():
    res, stats, _ = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07"])],
        [_dispo("Dupont Jean", {"2026-09-07": {"matin": True, "pm": False}})],
        _aff_simple("BTS MCO 73", "Jean Dupont"),
    )
    assert res["2026-09-07"]["BTS MCO 73"]["formateur"] == "Jean Dupont"
    assert stats['assigned'] == 1

@_p2("assigner: heures épuisées → ⚠️")
def _():
    aff = {"BTS MCO 73": [{'formateur': "Jean Dupont", 'matiere': "Gestion",
                            'heures': 100, 'heures_faites': 100, 'priorite': 1}]}
    res, stats, _ = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07"])],
        [_dispo("Jean Dupont", {"2026-09-07": {"matin": True, "pm": False}})],
        aff,
    )
    assert res["2026-09-07"]["BTS MCO 73"]["formateur"] == "⚠️"

@_p2("assigner: plusieurs classes indépendantes")
def _():
    aff = {
        "BTS MCO 73": [{'formateur': "Dupont", 'matiere': "Gestion",
                         'heures': 100, 'heures_faites': 0, 'priorite': 1}],
        "BAC PRO 28": [{'formateur': "Martin", 'matiere': "Vente",
                         'heures': 80, 'heures_faites': 0, 'priorite': 1}],
    }
    res, _, _ = _app.assigner(
        [_planning("BTS MCO 73", ["2026-09-07"]), _planning("BAC PRO 28", ["2026-09-07"])],
        [_dispo("Dupont", {"2026-09-07": {"matin": True, "pm": True}}),
         _dispo("Martin", {"2026-09-07": {"matin": True, "pm": True}})],
        aff,
    )
    assert res["2026-09-07"]["BTS MCO 73"]["formateur"] == "Dupont"
    assert res["2026-09-07"]["BAC PRO 28"]["formateur"] == "Martin"


# Routes Flask ------------------------------------------------------------

def _make_planning_file_bytes():
    wb = _make_planning_wb([(2026, 9, 7)])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def _make_dispo_file_bytes():
    wb = _make_dispo_wb("Test Formateur", [(2026, 9, 7, True, False)])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


@_p2("route: GET / → 200")
def _():
    client = _app.app.test_client()
    r = client.get("/")
    assert r.status_code == 200

@_p2("route: GET /ping → 200 + 'ok'")
def _():
    client = _app.app.test_client()
    r = client.get("/ping")
    assert r.status_code == 200
    assert b"ok" in r.data.lower()

@_p2("route: POST /generer-template-vierge → job_id ou fichier")
def _():
    import json
    _app.app.config['TESTING'] = True
    client = _app.app.test_client()
    r = client.post("/generer-template-vierge",
                    data={"mois_json": '[[2026, 9]]'},
                    content_type="multipart/form-data")
    assert r.status_code in (200, 202), f"status={r.status_code}"
    if r.content_type and 'json' in r.content_type:
        body = json.loads(r.data)
        assert 'job_id' in body

@_p2("route: POST /generer-template sans fichier → erreur")
def _():
    _app.app.config['TESTING'] = True
    client = _app.app.test_client()
    r = client.post("/generer-template", data={}, content_type="multipart/form-data")
    assert r.status_code in (400, 500), f"status={r.status_code}"

@_p2("route: POST /generer-template avec planning → non-erreur")
def _():
    _app.app.config['TESTING'] = True
    client = _app.app.test_client()
    r = client.post("/generer-template",
                    data={"planning_0": (_make_planning_file_bytes(), "planning.xlsx"),
                          "mois_json": "[[2026, 9]]"},
                    content_type="multipart/form-data")
    assert r.status_code in (200, 400), f"status={r.status_code}"

@_p2("route: POST /generer sans fichier → erreur")
def _():
    _app.app.config['TESTING'] = True
    client = _app.app.test_client()
    r = client.post("/generer", data={}, content_type="multipart/form-data")
    assert r.status_code in (400, 500), f"status={r.status_code}"


# ── Runner partie 2 ──────────────────────────────────────────────────────────

def run_parsers():
    if not _APP_OK:
        print("\n[PARTIE 2 IGNORÉE] app.py non importable\n")
        return True

    total = passed = failed = 0
    errors = []

    print(f"\n{'Test':<55} {'Résultat'}")
    print("-" * 65)

    for name, fn in P2_TESTS:
        total += 1
        try:
            fn()
            passed += 1
            print(f"{name:<55} OK")
        except Exception as e:
            failed += 1
            errors.append(f"  [{name}] {e}")
            print(f"{name:<55} ECHEC")

    print("-" * 65)
    print(f"\n{total} tests | {passed} passés", end="")
    if failed:
        print(f" | {failed} échec(s)\n")
        for e in errors:
            print(e)
    else:
        print(" — TOUT OK")

    return failed == 0


if __name__ == '__main__':
    ok1 = run_all()
    ok2 = run_parsers()
    import sys as _sys
    _sys.exit(0 if (ok1 and ok2) else 1)
